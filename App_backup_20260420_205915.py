# -*- coding: utf-8 -*-
"""
Ad Campaign Report Generator - Flask + openpyxl
Run: pip install flask openpyxl pandas werkzeug && python App.py
"""

import os
import re
import math
import uuid
import time
import random
import threading
from pathlib import Path
from datetime import datetime, timezone

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.errors import IgnoredError
from flask import Flask, request, jsonify, send_file, after_this_request
from werkzeug.utils import secure_filename

# -- Config -------------------------------------------------------------------
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
jobs: dict[str, dict] = {}

APP_DB_FILE = "App_Url Data base.xlsx"
_DB_LOCK = threading.Lock()

# -- Excel styling helpers -----------------------------------------------------
HEADER_BG = "00B0F0"
HEADER_FG = "000000"
TOTAL_BG  = "9BC2E6"
TOTAL_FG  = "000000"

def _thin_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(ws, row, num_cols):
    fill = PatternFill("solid", fgColor=HEADER_BG)
    font = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 14.50

def style_data(ws, start_row, end_row, num_cols):
    font = Font(size=10, name="Calibri")
    for r in range(start_row, end_row):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.border = _thin_border()

def style_total_row(ws, row, num_cols):
    fill = PatternFill("solid", fgColor=TOTAL_BG)
    font = Font(bold=True, color=TOTAL_FG, size=11, name="Calibri")
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_fit(ws):
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 50)

def write_sheet(ws, headers, rows, total_row=None, formulas=None, alignments=None):
    ws.sheet_view.showGridLines = True
    formulas = formulas or {}
    alignments = alignments or {}
    col_map = {h: get_column_letter(i) for i, h in enumerate(headers, 1)}

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    def _assign(cell, val, h_name=None, row_idx=None):
        if hasattr(val, "item"):
            val = val.item()
        if h_name in formulas and row_idx is not None:
            f_temp = formulas[h_name]
            for h_ref, let in col_map.items():
                f_temp = f_temp.replace(f"{{{h_ref}}}", f"{let}{row_idx}")
            cell.value = f"={f_temp}"
            if h_name in ["CTR", "Click Rate (CTR)", "Viewability", "VCR (Completion Rate)"]:
                cell.number_format = "0.00%"
            return
        if isinstance(val, str) and val.strip().endswith("%"):
            try:
                cell.value = float(val.strip()[:-1]) / 100.0
                cell.number_format = "0.00%" if "." in val else "0%"
            except Exception:
                cell.value = val
        elif isinstance(val, str) and val.lstrip('-').isdigit():
            cell.value = int(val)
        else:
            cell.value = val

    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            _assign(cell, row.get(h, ""), h_name=h, row_idx=r_idx)
            target_align = "center"
            if isinstance(alignments, str):
                target_align = alignments
            elif h in alignments:
                target_align = alignments[h]
            elif h == headers[0] and not alignments:
                target_align = "left"
            elif not alignments:
                target_align = "right"
            cell.alignment = Alignment(horizontal=target_align, vertical="center")

    style_data(ws, 2, 2 + len(rows), len(headers))

    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = _thin_border()

    # ── Grand Total row ───────────────────────────────────────────────────────
    # Rules (in priority order):
    #   1. First column  → label text ("Grand Total", etc.)
    #   2. SUM_COLS      → always =SUM(X2:Xn)  — never hardcode numeric totals
    #   3. In formulas   → derived formula referencing the total row cells
    #                       (keeps CTR / Viewability / VCR live in Excel)
    #   4. Everything else → fallback plain value from total_row dict
    # ─────────────────────────────────────────────────────────────────────────
    if total_row:
        t_row = 2 + len(rows)
        SUM_COLS = {
            "Impressions", "Clicks", "Reach",
            "Measurable Impressions", "Viewable Impressions",
            "Sum of Starts (Video)", "Sum of Complete Views (Video)",
        }
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=t_row, column=c_idx)
            h_name = headers[c_idx - 1]

            if c_idx == 1:
                # First column is always the label
                _assign(cell, total_row.get(h_name, "Grand Total"))

            elif h_name in SUM_COLS:
                # Always use a SUM formula — never hardcode accumulation columns
                let = col_map[h_name]
                cell.value = f"=SUM({let}2:{let}{t_row-1})"

            elif h_name in formulas:
                # Derived rate / ratio: reference THIS total row so it stays live
                # e.g. CTR total = total_clicks_cell / total_impressions_cell
                f_temp = formulas[h_name]
                for h_ref, let in col_map.items():
                    f_temp = f_temp.replace(f"{{{h_ref}}}", f"{let}{t_row}")
                cell.value = f"={f_temp}"
                cell.number_format = "0.00%"

            else:
                # Any remaining column (plain label override, etc.)
                _assign(cell, total_row.get(h_name, ""))

        style_total_row(ws, t_row, len(headers))

    auto_fit(ws)

    try:
        ws.ignored_errors.append(IgnoredError(sqref="A1:ZZ5000",
                                              numberStoredAsText=True,
                                              formulaOmitsAdjacentData=True,
                                              inconsistentFormula=True,
                                              evalError=True,
                                              emptyCellReference=True,
                                              calculatedColumn=True))
    except Exception:
        pass


# -- JS-Python utility functions -----------------------------------------------
def rand_float(mn, mx):
    return mn + random.random() * (mx - mn)

def rand_int(mn, mx):
    return random.randint(mn, mx)

def pct(num, den, decimals=2):
    if den == 0:
        return "0.00%"
    return f"{round((num / den) * 100, decimals):.{decimals}f}%"

def safe_float(val, default=0.0):
    if val is None:
        return default
    s = str(val).strip().replace(',', '').replace(' ', '')
    if s.endswith('%'):
        try:
            return float(s[:-1]) / 100.0
        except ValueError:
            return default
    try:
        return float(s)
    except ValueError:
        return default

def safe_int(val, default=0):
    return int(safe_float(val, default))

def serial_to_date(serial):
    try:
        serial = float(serial)
        utc_days = int(serial - 25569)
        dt = datetime(1970, 1, 1, tzinfo=timezone.utc) + \
             __import__('datetime').timedelta(days=utc_days)
        return dt.strftime("%d %B, %Y")
    except Exception:
        return str(serial)


def deduplicate_preserving_sum(values: list[int], gap: int = 1) -> list[int]:
    _n = len(values)
    if _n < 2: return values
    items = [[i, v] for i, v in enumerate(values)]
    items.sort(key=lambda x: x[1], reverse=True)
    used: set[int] = set()

    for k in range(_n):
        val = items[k][1]
        if val not in used and all(abs(val - u) >= gap for u in used):
            used.add(val); continue
        found = False
        for _gap in sorted([gap, 5, 3, 2, 1], reverse=True):
            if _gap > gap: continue
            for _ in range(50):
                off  = random.randint(_gap, max(_gap + 5, _n))
                sign = random.choice([-1, 1])
                cand = val + sign * off
                if cand >= 1 and cand not in used and all(abs(cand - u) >= _gap for u in used):
                    items[k][1] = cand; used.add(cand); found = True; break
            if found: break
        if not found:
            off = 1
            while True:
                for cand in [val + off, val - off]:
                    if cand >= 1 and cand not in used:
                        items[k][1] = cand; used.add(cand); found = True; break
                if found: break
                off += 1

    drift = sum(values) - sum(x[1] for x in items)
    if drift:
        items.sort(key=lambda x: x[1], reverse=True)
        for k in range(_n):
            if not drift: break
            step = 1 if drift > 0 else -1
            cand = items[k][1] + step
            if cand >= 1 and cand not in used:
                items[k][1] = cand; used.add(cand); drift -= step
        if drift:
            items[0][1] += drift
    items.sort(key=lambda x: x[0])
    return [x[1] for x in items]

# -- URL cleaning utility (module-level) ---------------------------------------
def _clean_url(u: str) -> str:
    u = str(u).lower().strip()
    for prefix in ("https://", "http://", "www."):
        if u.startswith(prefix):
            u = u[len(prefix):]
    return u.rstrip("/")


# -- Line Item column auto-detection ------------------------------------------
LINE_ITEM_CANDIDATES = [
    # Ordered by priority — most specific/reliable first.
    # ID columns intentionally excluded (they hold codes, not names).
    "line item name",
    "line item",
    "lineitem",
    "line_item",
    "campaign line item",
    "ad name",
    "ad set name",
    "adset name",
    "placement",
    "ad unit",
]

# Columns that sound like line items but are actually ID/code columns — skip these.
_LI_EXCLUDE_KEYWORDS = {"id", "code", "key", "number", "num", "no"}

def detect_line_item_col(df: pd.DataFrame) -> str | None:
    """
    Return the column most likely to contain line item NAMES (not IDs).
    Prefers 'Line Item Name' > 'Line Item' > other candidates.
    Excludes columns that look like ID/code columns.
    """
    cols_lower = {c.strip().lower(): c for c in df.columns}

    # Pass 1: exact match against priority list
    for candidate in LINE_ITEM_CANDIDATES:
        if candidate in cols_lower:
            return cols_lower[candidate]

    # Pass 2: fuzzy — column contains "line item" but NOT an id/code word
    for lower, original in cols_lower.items():
        if "line" in lower and "item" in lower:
            tokens = set(re.split(r'[\s_\-]+', lower))
            if not tokens & _LI_EXCLUDE_KEYWORDS:   # skip if looks like an ID col
                return original

    return None


# -- Language detection from Line Item values ----------------------------------
LANGUAGE_KEYWORDS = {
    "cantonese":  "Cantonese",
    "mandarin":   "Mandarin",
    "chinese":    "Chinese",
    "hong kong":  "Cantonese",
    "hk":         "Cantonese",
    "taiwan":     "Mandarin",
    "tw":         "Mandarin",
    "arabic":     "Arabic",
    "english":    "English",
    "french":     "French",
    "spanish":    "Spanish",
    "hindi":      "Hindi",
    "malay":      "Malay",
    "tamil":      "Tamil",
    "tagalog":    "Filipino/Tagalog",
    "filipino":   "Filipino/Tagalog",
    "indonesian": "Indonesian",
    "thai":       "Thai",
    "vietnamese": "Vietnamese",
    "korean":     "Korean",
    "japanese":   "Japanese",
    "italian":    "Italian",
    "urdu":       "Urdu",
    "bengali":    "Bengali",
    "persian":    "Dari/Persian",
    "dari":       "Dari/Persian",
    "turkish":    "Turkish",
    "greek":      "Greek",
    "punjabi":    "Punjabi",
    "gujarati":   "Gujarati",
    "malayalam":  "Malayalam",
    "telugu":     "Telugu",
    "sinhala":    "Sinhala/Tamil",
    "khmer":      "Khmer",
    "nepali":     "Nepali",
    "amharic":    "Amharic",
    "burmese":    "Burmese/Karen",
    "karen":      "Burmese/Karen",
    "fijian":     "Fijian",
    "dinka":      "Dinka",
    "croatian":   "Croatian",
    "macedonian": "Macedonian",
    "assyrian":   "Assyrian/Arabic",
    "portuguese": "Spanish",
    "german":     "Multi-language",
    "russian":    "Multi-language",
}

_DB_MASTER_SHEET  = "Sheet1"
_DB_URL_COL       = "URL / App Name"
_DB_LANG_COL      = "Language or Line Item"
_DB_ID_COL        = "ID"
_DB_HEADER_ROW    = 2   # 1-indexed, row where column headers live
_DB_JUNK          = {"nan", "none", "", "app/url", "app", "url", "url / app name",
                     "language or line item"}

# ── Curated master list of languages the UI should always offer (order matters).
# DB languages NOT in this list are still loaded dynamically and appended. ──
CURATED_LANGUAGES = [
    "Chinese", "Arabic", "Vietnamese", "Korean", "Japanese", "Turkish", "Spanish",
    "Singaporean", "Malay", "Sudanesh", "Indonesian", "Filipino OR Tagalog",
    "Thai", "Taiwan", "Hong kong", "Italian", "French", "African", "Greek",
    "Mecedonians", "Croatians", "Nepalese", "(Dari) OR (Persian) OR (Farsi)",
    "Chin Haka", "Dinka", "Fiji", "Combodian OR Khmer", "Amharic",
    "(Burmash) OR (Karen)", "Assyrian", "South East Asian",
    "(African American) OR (African)", "Nigerians", "Ghanaians", "Egyptians",
    "Kenyans", "South Africa", "Moroccans", "Somalia", "Algerians", "Senegal",
    "Ethiopians", "Cameroons", "DR Congo", "Tanzania", "African other sites",
    "Pacific Islander", "New Zealand", "Samoa", "South Asian",
    "Indian-Tamil OR Tamil", "Indian Gujarati OT Gujarati",
    "Indian Malayalam OR Malayalam", "Indian Telugu - Telugu",
    "Indian_Punjabi OR Punjabi", "PUNJABI (CA) OR Punjabi Canada",
    "Pakistan", "Srilankan", "Bangaladesh", "Chinese_AU OR Chinese  Australia",
    "Chinesh_NZ OR Chinesh New Zealand", "Mandarin",
    "Mandarin CA OR Mandarin Canada", "Cantonese", "Cantonese Canada",
]


def _norm_lang(s: str) -> str:
    """Normalize a language string for matching (lowercase, collapse whitespace)."""
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _load_db_master() -> pd.DataFrame:
    """
    Load the Master DB as a flat DataFrame with the Language column forward-filled
    so that every URL row carries its owning Language. Junk/header rows are
    dropped. Read-only pandas view — does NOT preserve workbook formatting.
    """
    try:
        df = pd.read_excel(APP_DB_FILE, sheet_name=_DB_MASTER_SHEET,
                           header=_DB_HEADER_ROW - 1)
    except Exception:
        return pd.DataFrame()

    if _DB_LANG_COL not in df.columns or _DB_URL_COL not in df.columns:
        return pd.DataFrame()

    # Drop literal-header-text rows that repeat inside the sheet
    df = df[df[_DB_LANG_COL].astype(str).str.strip().str.lower()
            != _DB_LANG_COL.lower()]

    # Forward-fill language so every row inherits the section language
    df[_DB_LANG_COL] = df[_DB_LANG_COL].ffill()

    # Drop URL-less rows
    df = df[df[_DB_URL_COL].notna()]
    df = df[~df[_DB_URL_COL].astype(str).str.strip().str.lower().isin(_DB_JUNK)]
    df = df.reset_index(drop=True)
    return df


def _load_db_records() -> list[dict]:
    """Return list of {"language": str, "url": str} records from DB (clean)."""
    df = _load_db_master()
    out: list[dict] = []
    if df.empty:
        return out
    for _, row in df.iterrows():
        lang = str(row[_DB_LANG_COL]).strip()
        url  = str(row[_DB_URL_COL]).strip()
        if not lang or not url or _clean_url(url) in _DB_JUNK:
            continue
        out.append({"language": lang, "url": url})
    return out


def list_db_languages() -> list[str]:
    """All distinct language names currently in the DB, in first-seen order."""
    seen: set[str] = set()
    order: list[str] = []
    try:
        wb = openpyxl.load_workbook(APP_DB_FILE, read_only=True, data_only=True)
        ws = wb[_DB_MASTER_SHEET] if _DB_MASTER_SHEET in wb.sheetnames else wb.active
        # Find the language column index by the header row
        header_row = next(ws.iter_rows(min_row=_DB_HEADER_ROW,
                                        max_row=_DB_HEADER_ROW, values_only=True))
        try:
            lang_idx = header_row.index(_DB_LANG_COL)
        except ValueError:
            lang_idx = 2
        for row in ws.iter_rows(min_row=_DB_HEADER_ROW + 1, values_only=True):
            val = row[lang_idx] if lang_idx < len(row) else None
            if val is None:
                continue
            s = str(val).strip()
            if not s or s.lower() == _DB_LANG_COL.lower():
                continue
            key = _norm_lang(s)
            if key in seen:
                continue
            seen.add(key)
            order.append(s)
        wb.close()
    except Exception:
        pass
    return order


def list_languages_for_ui() -> list[str]:
    """Curated list first, then any DB-only languages appended (de-duped)."""
    db_langs = list_db_languages()
    seen_norm: dict[str, str] = {}
    ordered: list[str] = []
    for name in CURATED_LANGUAGES + db_langs:
        key = _norm_lang(name)
        if key and key not in seen_norm:
            seen_norm[key] = name
            ordered.append(name)
    return ordered


def _find_language_row_range(ws, language: str) -> tuple[int, int] | None:
    """
    Locate a language section in the worksheet and return (start_row, end_row)
    inclusive, where start_row is the first data row of that language and
    end_row is the last. Returns None if language not found.
    """
    target = _norm_lang(language)
    lang_col_idx = 3  # column C: "Language or Line Item"
    # First, forward-fill by scanning
    first = None
    last  = None
    current_lang = None
    for r in range(_DB_HEADER_ROW + 1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=lang_col_idx).value
        if cell_val is not None and str(cell_val).strip().lower() != _DB_LANG_COL.lower():
            current_lang = _norm_lang(cell_val)
        if current_lang == target:
            if first is None:
                first = r
            last = r
    if first is None:
        return None
    return first, last


def add_url_to_db(url: str, language: str) -> bool:
    """
    Append a URL under a language section in the DB. If the language section
    doesn't exist yet, the language is added (new section at the bottom).
    Returns True if something was added, False if URL already present under
    the language or on any failure.
    """
    if not url or not language:
        return False
    url       = str(url).strip()
    language  = str(language).strip()
    url_clean = _clean_url(url)
    if url_clean in _DB_JUNK:
        return False

    with _DB_LOCK:
        try:
            wb = openpyxl.load_workbook(APP_DB_FILE)
            ws = wb[_DB_MASTER_SHEET] if _DB_MASTER_SHEET in wb.sheetnames else wb.active

            # Check the language exists already — and check for duplicate URL
            rng = _find_language_row_range(ws, language)
            if rng is not None:
                start_r, end_r = rng
                for r in range(start_r, end_r + 1):
                    existing = ws.cell(row=r, column=2).value
                    if existing and _clean_url(str(existing)) == url_clean:
                        wb.close()
                        return False  # already present
                # Next ID in section
                last_id = 0
                for r in range(start_r, end_r + 1):
                    v = ws.cell(row=r, column=1).value
                    try:
                        last_id = max(last_id, int(v))
                    except Exception:
                        pass
                new_row = end_r + 1
                # Shift rows down below this section to keep adjacent sections intact
                ws.insert_rows(new_row)
                ws.cell(row=new_row, column=1, value=last_id + 1)
                ws.cell(row=new_row, column=2, value=url)
                ws.cell(row=new_row, column=3, value=None)  # blank — inherits
            else:
                # Append a new language section at the very end
                last_row = ws.max_row + 1
                ws.cell(row=last_row,     column=1, value=1)
                ws.cell(row=last_row,     column=2, value=url)
                ws.cell(row=last_row,     column=3, value=language)

            wb.save(APP_DB_FILE)
            wb.close()
            return True
        except Exception as e:
            print(f"[DB] add_url_to_db failed for {url} / {language}: {e}")
            return False


def add_language_to_db(language: str) -> bool:
    """Add an empty language section to the DB. No-op if already present."""
    language = str(language or "").strip()
    if not language:
        return False

    with _DB_LOCK:
        try:
            wb = openpyxl.load_workbook(APP_DB_FILE)
            ws = wb[_DB_MASTER_SHEET] if _DB_MASTER_SHEET in wb.sheetnames else wb.active
            if _find_language_row_range(ws, language) is not None:
                wb.close()
                return False  # already exists
            last_row = ws.max_row + 1
            # Seed with a placeholder blank URL row — section becomes navigable.
            ws.cell(row=last_row, column=1, value=1)
            ws.cell(row=last_row, column=2, value="")
            ws.cell(row=last_row, column=3, value=language)
            wb.save(APP_DB_FILE)
            wb.close()
            return True
        except Exception as e:
            print(f"[DB] add_language_to_db failed for {language}: {e}")
            return False


def remove_language_from_db(language: str) -> bool:
    """Remove a whole language section (all its rows) from the DB."""
    language = str(language or "").strip()
    if not language:
        return False

    with _DB_LOCK:
        try:
            wb = openpyxl.load_workbook(APP_DB_FILE)
            ws = wb[_DB_MASTER_SHEET] if _DB_MASTER_SHEET in wb.sheetnames else wb.active
            rng = _find_language_row_range(ws, language)
            if rng is None:
                wb.close()
                return False
            start_r, end_r = rng
            count = end_r - start_r + 1
            ws.delete_rows(start_r, count)
            wb.save(APP_DB_FILE)
            wb.close()
            return True
        except Exception as e:
            print(f"[DB] remove_language_from_db failed for {language}: {e}")
            return False


def detect_language_from_df1(df1: pd.DataFrame) -> str | None:
    target_col = None
    cols = [str(c).lower().strip() for c in df1.columns]
    for cand in ["campaign name", "campaign"]:
        if cand in cols:
            target_col = df1.columns[cols.index(cand)]; break
    if target_col is None:
        target_col = detect_line_item_col(df1)
    if target_col is None:
        return None
    combined_text = " ".join(
        df1[target_col].dropna().astype(str).str.strip().str.lower().unique()
    )
    for keyword, lang_name in LANGUAGE_KEYWORDS.items():
        if keyword in combined_text:
            return lang_name
    return None


def extract_language_from_line_item(line_item_str: str) -> str | None:
    text = str(line_item_str).strip()
    tokens = re.split(r'\s*[-_|:]+\s*|\s{2,}', text)
    for token in tokens:
        tl = token.strip().lower()
        if tl in LANGUAGE_KEYWORDS:
            return LANGUAGE_KEYWORDS[tl]
    text_lower = text.lower()
    for keyword, lang_name in LANGUAGE_KEYWORDS.items():
        if keyword in text_lower:
            return lang_name
    return None


def get_db_urls_by_priority(language: str | None) -> dict[str, list[str]]:
    df = _load_db_master()
    res = {"primary": [], "broad": [], "multi": [], "all": []}
    if df.empty or _DB_URL_COL not in df.columns:
        return res

    def _get_cleaned(mask):
        return [
            _clean_url(u) for u in df.loc[mask, _DB_URL_COL].dropna().astype(str).tolist()
            if _clean_url(u) not in _DB_JUNK
        ]

    seen = set()
    if _DB_LANG_COL in df.columns:
        ml_urls = _get_cleaned(df[_DB_LANG_COL].astype(str).str.strip().str.lower() == "multi-language")
        for u in ml_urls:
            if u not in seen: res["multi"].append(u); seen.add(u)

    if language and _DB_LANG_COL in df.columns:
        lang_lower = language.strip().lower()
        pri_urls = _get_cleaned(df[_DB_LANG_COL].astype(str).str.strip().str.lower() == lang_lower)
        for u in pri_urls:
            if u not in seen: res["primary"].append(u); seen.add(u)
        if lang_lower in ["cantonese", "mandarin"]:
            broad_urls = _get_cleaned(df[_DB_LANG_COL].astype(str).str.strip().str.lower() == "chinese")
            for u in broad_urls:
                if u not in seen: res["broad"].append(u); seen.add(u)

    all_urls = _get_cleaned(pd.Series([True] * len(df)))
    for u in all_urls:
        if u not in seen: res["all"].append(u); seen.add(u)

    return res


def load_db_urls(language: str | None) -> list[str]:
    p = get_db_urls_by_priority(language)
    return p["primary"] + p["broad"] + p["multi"] + p["all"]


def get_validated_urls_from_db(
    languages: list[str],
    user_urls: list[str],
) -> list[str]:
    if not languages or not user_urls:
        return []
    df = _load_db_master()
    if df.empty or _DB_URL_COL not in df.columns:
        return []

    lang_set   = {l.lower() for l in languages if l}
    user_clean = {_clean_url(u) for u in user_urls
                  if _clean_url(u) not in _DB_JUNK}
    if not user_clean:
        return []

    validated: list[str] = []
    seen: set[str] = set()
    has_lang_col = _DB_LANG_COL in df.columns

    for _, row in df.iterrows():
        url_val = _clean_url(str(row[_DB_URL_COL]))
        if url_val in _DB_JUNK or url_val not in user_clean:
            continue
        if has_lang_col:
            db_lang = str(row[_DB_LANG_COL]).strip().lower()
            if not any(lang in db_lang or db_lang in lang for lang in lang_set):
                continue
        if url_val not in seen:
            validated.append(url_val)
            seen.add(url_val)

    return validated


# -- Sheet builders ------------------------------------------------------------

def build_sheet1_reach(total_imp, total_clk):
    ctr = pct(total_clk, total_imp)
    return [{
        "Impressions": total_imp,
        "Clicks": total_clk,
        "Click Rate (CTR)": ctr,
        "Reach": "",
        "Frequency": ""
    }]


def build_sheet2_date(df, total_imp, total_clk, ctr_reach):
    rows = []
    sum_imp = sum_clk = sum_view = sum_meas = sum_starts = sum_comp = 0
    vcr_weighted = vcr_imp_total = 0

    for _, r in df.iterrows():
        imp   = safe_int(r.get("Impressions", 0))
        clk   = safe_int(r.get("Clicks", 0))
        view  = safe_int(r.get("Viewable Impressions", 0))
        meas  = safe_int(r.get("Measurable Impressions", 0))
        starts= safe_int(r.get("Start views", 0))
        comps = safe_int(r.get("Complete Views", 0))
        vcr_raw = safe_float(r.get("Video Completion Rate (VCR)", 0))

        date_val = r.get("Date", "")
        try:
            float(date_val)
            date_str = serial_to_date(date_val)
        except (ValueError, TypeError):
            date_str = str(date_val)

        sum_imp   += imp;   sum_clk   += clk
        sum_view  += view;  sum_meas  += meas
        sum_starts+= starts;sum_comp  += comps
        vcr_weighted   += vcr_raw * imp
        vcr_imp_total  += imp

        vcr_pct = f"{round(vcr_raw * 100)}%" if vcr_raw > 0 else "0%"

        rows.append({
            "Date": date_str,
            "Impressions": imp,
            "Clicks": clk,
            "Click Rate (CTR)": pct(clk, imp),
            "Viewable Impressions": view,
            "Measurable Impressions": meas,
            "Viewability": pct(view, meas),
            "Sum of Starts (Video)": starts,
            "Sum of Complete Views (Video)": comps,
            "VCR (Completion Rate)": vcr_pct
        })

    avg_vcr = f"{round((vcr_weighted / vcr_imp_total) * 100)}%" if vcr_imp_total > 0 else "0%"

    total = {
        "Date": "Grand Total",
        "Impressions": sum_imp,
        "Clicks": sum_clk,
        "Click Rate (CTR)": ctr_reach,
        "Viewable Impressions": sum_view,
        "Measurable Impressions": sum_meas,
        "Viewability": pct(sum_view, sum_meas),
        "Sum of Starts (Video)": sum_starts,
        "Sum of Complete Views (Video)": sum_comp,
        "VCR (Completion Rate)": avg_vcr
    }
    return rows, total


def build_sheet3_timeofday(total_imp, total_clk, ctr_reach):
    hour_weights = {
        0:0.5,1:0.4,2:0.3,3:0.3,4:0.4,5:0.7,6:1.2,7:1.5,
        8:2.0,9:2.5,10:2.3,11:2.4,12:2.2,13:1.8,14:1.6,15:1.7,
        16:1.9,17:2.6,18:2.8,19:2.5,20:2.0,21:1.6,22:1.2,23:0.8
    }

    noisy = [hour_weights[h] * (0.85 + random.random() * 0.30) for h in range(24)]
    total_w = sum(noisy)
    hourly_imp = [int((w / total_w) * total_imp) for w in noisy]
    diff = total_imp - sum(hourly_imp)
    for _ in range(abs(diff)):
        hourly_imp[random.randint(0, 23)] += 1 if diff > 0 else -1

    hourly_imp = deduplicate_preserving_sum(hourly_imp, gap=1)

    hourly_clk = []
    for imp in hourly_imp:
        if imp >= 180:
            c_min = math.ceil(imp * 0.0035)
            c_max = math.floor(imp * 0.0056)
            if c_min > c_max:
                hourly_clk.append(c_min)
            else:
                hourly_clk.append(random.randint(c_min, c_max))
        else:
            hourly_clk.append(0)

    current = sum(hourly_clk)
    diff_clk = total_clk - current
    eligible = [i for i, imp in enumerate(hourly_imp) if imp >= 180]

    iterations = 0
    while diff_clk != 0 and eligible and iterations < 5000:
        iterations += 1
        idx = random.choice(eligible)
        imp = hourly_imp[idx]
        clk = hourly_clk[idx]
        if diff_clk > 0:
            if (clk + 1) / imp <= 0.0056:
                hourly_clk[idx] += 1; diff_clk -= 1
        elif diff_clk < 0:
            if clk > 1 and (clk - 1) / imp >= 0.0035:
                hourly_clk[idx] -= 1; diff_clk += 1

    if diff_clk != 0 and eligible:
        for _ in range(abs(diff_clk)):
            safe_indices = []
            for i in eligible:
                if diff_clk > 0 and (hourly_clk[i] + 1) / hourly_imp[i] <= 0.0056:
                    safe_indices.append(i)
                elif diff_clk < 0 and hourly_clk[i] > 1 and (hourly_clk[i] - 1) / hourly_imp[i] >= 0.0035:
                    safe_indices.append(i)
            if not safe_indices: break
            idx = random.choice(safe_indices)
            if diff_clk > 0:
                hourly_clk[idx] += 1; diff_clk -= 1
            else:
                hourly_clk[idx] -= 1; diff_clk += 1

    rows = []
    for h in range(24):
        imp = hourly_imp[h]; clk = hourly_clk[h]
        rows.append({
            "Time of Day": h,
            "Impressions": imp,
            "Clicks": clk,
            "Click Rate (CTR)": pct(clk, imp)
        })

    total = {
        "Time of Day": "Grand Total",
        "Impressions": total_imp,
        "Clicks": total_clk,
        "Click Rate (CTR)": ctr_reach
    }
    return rows, total


def build_sheet4_age(total_imp, total_clk, ctr_reach):
    pct_1824 = rand_float(0.20, 0.25)
    pct_4554 = rand_float(0.09, 0.12)
    remaining = max(0.60, min(0.74, 1 - pct_1824 - pct_4554))
    split = rand_float(0.45, 0.55)
    pct_2534 = remaining * split
    pct_3544 = remaining - pct_2534

    imp_1824 = round(total_imp * pct_1824)
    imp_4554 = round(total_imp * pct_4554)
    rem_imp  = total_imp - imp_1824 - imp_4554
    imp_2534 = round(rem_imp * split)
    imp_3544 = rem_imp - imp_2534
    imp_2534 += total_imp - (imp_1824 + imp_2534 + imp_3544 + imp_4554)

    groups = [
        {"age": "18-24", "imp": imp_1824},
        {"age": "25-34", "imp": imp_2534},
        {"age": "35-44", "imp": imp_3544},
        {"age": "45-54", "imp": imp_4554},
    ]

    clicks = [round(g["imp"] * rand_float(0.0035, 0.0056)) if g["imp"] >= 180 else 0 for g in groups]
    c_sum = sum(clicks)
    if c_sum > 0:
        clicks = [round(c * total_clk / c_sum) if groups[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]

    drift = total_clk - sum(clicks)
    clicks[clicks.index(max(clicks))] += drift

    rows = [{
        "Age": g["age"],
        "Impressions": g["imp"],
        "Clicks": clicks[i],
        "Click Rate (CTR)": pct(clicks[i], g["imp"])
    } for i, g in enumerate(groups)]

    total = {"Age": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


def build_sheet5_gender(total_imp, total_clk, ctr_reach):
    female_pct = rand_float(0.30, 0.40)
    imp_female = round(total_imp * female_pct)
    imp_male   = total_imp - imp_female
    male_pct   = imp_male / total_imp

    if male_pct < 0.58:
        imp_male   = round(total_imp * rand_float(0.58, 0.61))
        imp_female = total_imp - imp_male
    elif male_pct > 0.65:
        imp_male   = round(total_imp * rand_float(0.62, 0.65))
        imp_female = total_imp - imp_male

    imp_male += total_imp - (imp_male + imp_female)

    groups = [{"gender": "Male", "imp": imp_male}, {"gender": "Female", "imp": imp_female}]
    clicks = [round(g["imp"] * rand_float(0.0035, 0.0056)) if g["imp"] >= 180 else 0 for g in groups]
    g_sum  = sum(clicks)
    if g_sum > 0:
        clicks = [round(c * total_clk / g_sum) if groups[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]

    drift = total_clk - sum(clicks)
    clicks[clicks.index(max(clicks))] += drift

    rows = [{
        "Gender": g["gender"],
        "Impressions": g["imp"],
        "Clicks": clicks[i],
        "Click Rate (CTR)": pct(clicks[i], g["imp"])
    } for i, g in enumerate(groups)]

    total = {"Gender": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


def build_sheet6_device(total_imp, total_clk, ctr_reach):
    tablet_pct  = rand_float(0.07, 0.10)
    desktop_shr = rand_float(0.45, 0.55)
    mobile_shr  = 1 - desktop_shr
    rem         = 1 - tablet_pct

    imp_tablet  = round(total_imp * tablet_pct)
    imp_desktop = round(total_imp * rem * desktop_shr)
    imp_mobile  = total_imp - imp_tablet - imp_desktop
    imp_mobile += total_imp - (imp_desktop + imp_mobile + imp_tablet)

    groups = [
        {"device": "Desktop",     "imp": imp_desktop},
        {"device": "Smart Phone", "imp": imp_mobile},
        {"device": "Tablet",      "imp": imp_tablet},
    ]

    clicks = [max(0, int(g["imp"] * rand_float(0.0035, 0.0056) + rand_float(-0.4, 0.4)))
              if g["imp"] >= 180 else 0 for g in groups]
    d_sum  = sum(clicks)
    if d_sum > 0:
        clicks = [int(c * total_clk / d_sum) if groups[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]

    drift = total_clk - sum(clicks)
    clicks[clicks.index(max(clicks))] += drift

    rows = [{
        "Device Type": g["device"],
        "Impressions": g["imp"],
        "Clicks": clicks[i],
        "Click Rate (CTR)": pct(clicks[i], g["imp"])
    } for i, g in enumerate(groups)]

    total = {"Device Type": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


def build_sheet7_exchange(total_imp, total_clk, ctr_reach):
    tier1_other = ["BidSwitch","Index Exchange","Magnite DV+","PubMatic"]
    tier2 = ["Criteo Commerce Grid","Equativ","InMobi","Media.net","Nexxen (fka Unruly)","OpenX","Sovrn"]
    tier3 = ["Microsoft Monetize","Epsilon Core Private Exchange","Yieldmo","TripleLift",
             "Kargo","Improve Digital","TeadsTv","GumGum","Adform"]

    ex_t1_imp = int(total_imp * rand_float(0.80, 0.85))
    ex_t2_imp = int(total_imp * rand_float(0.08, 0.13))
    ex_t3_imp = total_imp - ex_t1_imp - ex_t2_imp

    google_imp   = int(ex_t1_imp * rand_float(0.35, 0.45))
    t1_other_imp = ex_t1_imp - google_imp

    def split_across(names, total):
        remaining = total
        result = []
        for i, name in enumerate(names):
            if i == len(names) - 1:
                imp = remaining
            else:
                imp = max(1, rand_int(1, int((remaining - (len(names) - 1 - i)) * 0.6)))
            remaining -= imp
            result.append({"name": name, "imp": imp})
        return result

    exchanges = (
        [{"name": "Google Ad Manager", "imp": google_imp}]
        + split_across(tier1_other, t1_other_imp)
        + split_across(tier2, ex_t2_imp)
        + split_across(tier3, ex_t3_imp)
    )

    drift = total_imp - sum(e["imp"] for e in exchanges)
    exchanges[0]["imp"] += drift

    clicks = [round(e["imp"] * rand_float(0.0035, 0.0056)) if e["imp"] >= 180 else 0 for e in exchanges]
    e_sum  = sum(clicks)
    if e_sum > 0:
        clicks = [round(c * total_clk / e_sum) if exchanges[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]

    e_drift = total_clk - sum(clicks)
    max_idx = max(range(len(exchanges)), key=lambda i: exchanges[i]["imp"])
    clicks[max_idx] += e_drift

    rows = sorted([{
        "Exchange": e["name"],
        "Impressions": e["imp"],
        "Clicks": clicks[i],
        "Click Rate (CTR)": pct(clicks[i], e["imp"])
    } for i, e in enumerate(exchanges)], key=lambda x: x["Exchange"])

    total = {"Exchange": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


# -- Sheet 8 - City -----------------------------------------------------------

def _normalize_text(val: str) -> str:
    val = str(val).strip()
    val = re.sub(r'[\u2013\u2014\u2212\u2015]', '-', val)
    val = re.sub(r'\s+', ' ', val)
    return val.lower().strip()


def _clean_li2(val: str) -> str:
    val = str(val).strip()
    val = re.sub(r'^[\w\-]+\s*\|\s*', '', val)
    return _normalize_text(val)


def _normalize_city(val: str) -> str:
    return re.sub(r'\s+', ' ', str(val).strip()).title()


def build_sheet8_city(df1: pd.DataFrame, df2: pd.DataFrame,
                      total_imp: int, total_clk: int, ctr_reach: str):
    info = {
        "matched_line_items": 0,
        "cities_found": 0,
        "warnings": [],
        "debug": {}
    }

    li_col1 = detect_line_item_col(df1)
    li_col2 = detect_line_item_col(df2)

    info["debug"]["file1_columns"]    = list(df1.columns)
    info["debug"]["file2_columns"]    = list(df2.columns)
    info["debug"]["li_col1_detected"] = li_col1
    info["debug"]["li_col2_detected"] = li_col2
    info["debug"]["file1_row_count"]  = len(df1)
    info["debug"]["file2_row_count"]  = len(df2)

    matched_df2 = pd.DataFrame()

    if li_col1 and li_col2:
        items1 = {
            _normalize_text(v)
            for v in df1[li_col1].dropna().unique()
            if str(v).strip()
        }
        items1.discard("")

        info["debug"]["items1_sample"] = list(items1)[:10]
        info["debug"]["items2_raw_sample"] = [
            str(v).strip() for v in df2[li_col2].dropna().unique()[:10]
        ]

        def matches_any(v2_raw: str) -> bool:
            cleaned = _clean_li2(v2_raw)
            return bool(cleaned) and cleaned in items1

        mask = df2[li_col2].apply(matches_any)
        matched_df2 = df2[mask].copy()

        info["matched_line_items"]   = int(mask.sum())
        info["debug"]["match_count"] = int(mask.sum())

        if matched_df2.empty:
            info["warnings"].append(
                f"Core matching found 0 rows "
                f"(File1 col='{li_col1}', File2 col='{li_col2}'). "
                "Falling back to all File 2 rows."
            )
    else:
        if li_col1 is None:
            info["warnings"].append(
                f"No Line Item column in File 1 "
                f"(cols: {list(df1.columns)[:8]}). Using all File 2 rows."
            )
        if li_col2 is None:
            info["warnings"].append(
                f"No Line Item column in File 2 "
                f"(cols: {list(df2.columns)[:8]}). Using all File 2 rows."
            )

    if matched_df2.empty:
        matched_df2 = df2.copy()
        info["matched_line_items"] = len(df2)
        info["debug"]["fallback_used"] = True
    else:
        info["debug"]["fallback_used"] = False

    city_col = None
    for col in matched_df2.columns:
        if col.strip().lower() == "city":
            city_col = col
            break
    if city_col is None:
        for col in matched_df2.columns:
            if "city" in col.strip().lower():
                city_col = col
                break

    info["debug"]["city_col_detected"] = city_col

    if city_col is None:
        info["warnings"].append(
            f"No 'City' column in File 2. "
            f"Columns found: {list(df2.columns)}. Sheet 8 will only show Grand Total."
        )
        return [], {
            "City": "Grand Total",
            "Impressions": total_imp,
            "Clicks": total_clk,
            "Click Rate (CTR)": ctr_reach
        }, info

    imp_col2 = None
    for col in matched_df2.columns:
        if col.strip().lower() == "impressions":
            imp_col2 = col
            break
    if imp_col2 is None:
        for col in matched_df2.columns:
            if "impression" in col.strip().lower():
                imp_col2 = col
                break

    info["debug"]["imp_col2_detected"] = imp_col2

    matched_df2 = matched_df2.copy()
    matched_df2["_city_norm"] = matched_df2[city_col].apply(_normalize_city)

    if li_col2:
        matched_df2["_li_clean"] = matched_df2[li_col2].apply(_clean_li2)
        matched_df2 = matched_df2.drop_duplicates(
            subset=["_li_clean", "_city_norm"]
        ).copy()

    if imp_col2:
        agg_df = (
            matched_df2
            .assign(_imp=matched_df2[imp_col2].apply(safe_float))
            .query("_city_norm.str.lower() not in ('nan', 'none', '', 'unknown')")
            .groupby("_city_norm", sort=False)["_imp"]
            .sum()
            .reset_index()
        )
        city_weight_map = dict(zip(agg_df["_city_norm"], agg_df["_imp"]))
    else:
        info["warnings"].append(
            "No Impressions column in File 2. Equal weights used for city distribution."
        )
        valid = (
            matched_df2["_city_norm"]
            .dropna()
            .astype(str)
            .str.strip()
        )
        city_weight_map = {
            c: 1.0
            for c in dict.fromkeys(valid)
            if c and c.lower() not in ("nan", "none", "", "unknown")
        }

    unique_cities = list(city_weight_map.keys())

    info["debug"]["city_agg_sample"]    = {k: v for k, v in list(city_weight_map.items())[:10]}
    info["debug"]["unique_city_count"]  = len(unique_cities)
    info["debug"]["unique_city_sample"] = unique_cities[:10]

    if not unique_cities:
        info["warnings"].append(
            f"City column '{city_col}' has no valid values after filtering. "
            "Sheet 8 will only show Grand Total."
        )
        return [], {
            "City": "Grand Total",
            "Impressions": total_imp,
            "Clicks": total_clk,
            "Click Rate (CTR)": ctr_reach
        }, info

    info["cities_found"] = len(unique_cities)
    n = len(unique_cities)

    weights = [max(float(city_weight_map[c]), 1.0) for c in unique_cities]
    w_total = sum(weights)

    info["debug"]["city_weights_sample"] = dict(zip(unique_cities[:10], weights[:10]))
    info["debug"]["w_total"] = w_total

    def largest_remainder(weights, w_total, total):
        exact   = [(w / w_total) * total for w in weights]
        floored = [int(e) for e in exact]
        fracs   = [e - f for e, f in zip(exact, floored)]
        remainder = total - sum(floored)
        order = sorted(range(len(weights)), key=lambda i: fracs[i], reverse=True)
        for i in range(remainder):
            floored[order[i]] += 1
        return floored

    city_imps = largest_remainder(weights, w_total, total_imp)

    def _dedup_city_imps(values):
        n = len(values)
        if sum(values) < n * (n + 1) // 2:
            return values
        items = [[i, v] for i, v in enumerate(values)]
        items.sort(key=lambda x: x[1], reverse=True)
        used = set()
        for k in range(n):
            val = items[k][1]
            if val in used:
                offset = 1
                new_val = val
                while True:
                    if val - offset >= 1 and (val - offset) not in used:
                        new_val = val - offset; break
                    if (val + offset) not in used:
                        new_val = val + offset; break
                    offset += 1
                items[k][1] = new_val
                used.add(new_val)
            else:
                used.add(val)
        current_sum = sum(x[1] for x in items)
        drift = sum(values) - current_sum
        if drift > 0:
            max_item = max(items, key=lambda x: x[1])
            max_item[1] += drift
        elif drift < 0:
            while drift < 0:
                items.sort(key=lambda x: x[1], reverse=True)
                subtracted = False
                for x in items:
                    if x[1] - 1 >= 1 and (x[1] - 1) not in used:
                        used.remove(x[1])
                        x[1] -= 1
                        used.add(x[1])
                        drift += 1
                        subtracted = True
                        break
                if not subtracted:
                    break
        items.sort(key=lambda x: x[0])
        return [x[1] for x in items]

    city_imps = _dedup_city_imps(city_imps)

    raw_clicks = [
        max(1, int(imp * rand_float(0.35/100, 0.56/100))) if imp >= 180 else 0
        for imp in city_imps
    ]
    sum_raw = sum(raw_clicks)
    if sum_raw > 0:
        city_clks = [max(1, int(round(r * total_clk / sum_raw))) if city_imps[i] >= 180 else 0 for i, r in enumerate(raw_clicks)]
    else:
        city_clks = [0] * len(city_imps)

    drift = total_clk - sum(city_clks)
    if city_imps and drift != 0:
        eligible = [i for i, imp in enumerate(city_imps) if imp >= 180]
        if not eligible:
            eligible = list(range(len(city_imps)))
        if drift > 0:
            max_idx = max(eligible, key=lambda i: city_imps[i])
            city_clks[max_idx] += drift
        elif drift < 0:
            for _ in range(abs(drift)):
                valid = [i for i in eligible if city_clks[i] > 1]
                if not valid:
                    break
                max_clk_idx = max(valid, key=lambda i: city_clks[i])
                city_clks[max_clk_idx] -= 1

    rows = [
        {
            "City": city,
            "Impressions": imp,
            "Clicks": clk,
            "Click Rate (CTR)": pct(clk, imp)
        }
        for city, imp, clk in zip(unique_cities, city_imps, city_clks)
    ]
    rows.sort(key=lambda x: str(x["City"]).strip().lower())

    total_row = {
        "City": "Grand Total",
        "Impressions": total_imp,
        "Clicks": total_clk,
        "Click Rate (CTR)": ctr_reach
    }

    return rows, total_row, info


# -- Sheet 9 - Creative --------------------------------------------------------

def build_sheet9_creative(df1: pd.DataFrame, df2: pd.DataFrame,
                          total_imp: int, total_clk: int):
    info = {"warnings": []}
    li_col1 = detect_line_item_col(df1)

    creative_col = None
    for cand in ["creative", "ad name", "ad content", "ad"]:
        for col in df2.columns:
            if cand == str(col).strip().lower():
                creative_col = col; break
        if creative_col: break

    if not creative_col:
        for col in df2.columns:
            cl = str(col).strip().lower()
            if "creative" in cl and "id" not in cl:
                creative_col = col; break

    if not creative_col:
        for cand in ["creative", "ad"]:
            for col in df2.columns:
                if cand in str(col).strip().lower():
                    creative_col = col; break
            if creative_col: break

    if not creative_col:
        creative_col = detect_line_item_col(df2)

    if not creative_col:
        return [], {"Creative": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": pct(total_clk, total_imp)}, info

    matched_df2 = pd.DataFrame()
    if li_col1:
        items1 = [str(v).strip() for v in df1[li_col1].dropna().unique() if str(v).strip()]

        def get_tokens(s):
            return set(re.findall(r'[a-z0-9]+', s.lower()))

        def matches_any_creative(v2_raw):
            v2_tokens = get_tokens(str(v2_raw))
            if not v2_tokens: return False
            for i1 in items1:
                i1_tokens = get_tokens(i1)
                if not i1_tokens: continue
                if i1_tokens.issubset(v2_tokens) or v2_tokens.issubset(i1_tokens):
                    return True
            return False

        mask = df2[creative_col].apply(matches_any_creative)
        matched_df2 = df2[mask].copy()

    if matched_df2.empty:
        info["warnings"].append("No creative-based matches found for the provided Line Items.")
        return [], {"Creative": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": pct(total_clk, total_imp)}, info

    imp_col2 = clk_col2 = None
    for col in matched_df2.columns:
        cl = col.strip().lower()
        if "impression" in cl: imp_col2 = col
        if "click" in cl: clk_col2 = col

    agg_df = (
        matched_df2
        .assign(
            _creative=matched_df2[creative_col].astype(str).str.strip(),
            _imp=matched_df2[imp_col2].apply(safe_float) if imp_col2 else 1.0,
            _clk=matched_df2[clk_col2].apply(safe_float) if clk_col2 else 1.0
        )
        .query("_creative.str.lower() not in ('nan', 'none', '', 'unknown')")
        .groupby("_creative", sort=False)[["_imp", "_clk"]]
        .sum()
        .reset_index()
    )

    unique_creatives = list(agg_df["_creative"])
    if not unique_creatives:
        return [], {"Creative": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": pct(total_clk, total_imp)}, info

    imp_weights = [max(float(w), 1.0) for w in agg_df["_imp"]]
    clk_weights = [max(float(w), 0.0) for w in agg_df["_clk"]]
    w_imp_total = sum(imp_weights)
    if sum(clk_weights) == 0:
        clk_weights = imp_weights
    w_clk_total = sum(clk_weights)

    def largest_remainder(weights, w_total, total):
        if w_total == 0: w_total = 1.0
        exact = [(w / w_total) * total for w in weights]
        floored = [int(e) for e in exact]
        fracs = [e - f for e, f in zip(exact, floored)]
        remainder = int(total - sum(floored))
        order = sorted(range(len(weights)), key=lambda i: fracs[i], reverse=True)
        for i in range(remainder):
            floored[order[i]] += 1
        return floored

    crea_imps = largest_remainder(imp_weights, w_imp_total, total_imp)

    overall_ctr = total_clk / total_imp if total_imp > 0 else 0
    min_ctr = max(0, overall_ctr - 0.0014)
    max_ctr = overall_ctr + 0.0014

    raw_clks = [max(0.01, imp * rand_float(min_ctr, max_ctr)) if imp >= 180 else 0.0 for imp in crea_imps]
    w_clk_total = sum(raw_clks)
    if w_clk_total == 0:
        raw_clks = [1.0 if imp >= 180 else 0.0 for imp in crea_imps]
        w_clk_total = sum(raw_clks)

    crea_clks = largest_remainder(raw_clks, w_clk_total, total_clk)
    rows = [
        {"Creative": c, "Impressions": imp, "Clicks": clk, "Click Rate (CTR)": pct(clk, imp)}
        for c, imp, clk in zip(unique_creatives, crea_imps, crea_clks)
    ]
    rows.sort(key=lambda x: str(x["Creative"]).strip().lower())

    total_row = {
        "Creative": "Grand Total",
        "Impressions": total_imp,
        "Clicks": total_clk,
        "Click Rate (CTR)": pct(total_clk, total_imp)
    }
    return rows, total_row, info


# -- Sheet 10 - Apps / URLs ----------------------------------------------------

def extract_pivot_urls(df1: pd.DataFrame, df2: pd.DataFrame) -> list[str]:
    li_col1 = detect_line_item_col(df1)
    li_col2 = detect_line_item_col(df2)

    if not li_col1 or not li_col2:
        return []

    items1 = [str(v).strip().lower() for v in df1[li_col1].dropna().unique() if str(v).strip()]
    if not items1:
        return []

    items1_set = set(items1)
    def is_match(v):
        v_low = str(v).lower()
        for i in items1_set:
            if i in v_low: return True
        return False

    mask = df2[li_col2].apply(is_match)
    matched_df2 = df2[mask].copy()

    if matched_df2.empty:
        return []

    url_col = None
    for col in matched_df2.columns:
        cl = str(col).strip().lower()
        if cl in ["app/url", "url", "site", "domain", "app", "app url", "bundle id", "website", "apps"]:
            url_col = col; break

    if not url_col:
        for col in matched_df2.columns:
            cl = str(col).strip().lower()
            if any(k in cl for k in ["url", "site", "domain", "app", "bundle", "website"]):
                url_col = col; break

    if not url_col:
        for col in matched_df2.columns:
            if col != li_col2:
                url_col = col; break

    if not url_col:
        return []

    pivot_urls = (
        matched_df2[url_col]
        .dropna()
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    )
    pivot_urls = [u for u in pivot_urls if u.lower() not in ["nan", "none", "", "app", "url", "site", "domain"]]
    return pivot_urls


def build_sheet10_apps(
    app_urls_str: str,
    total_imp: int,
    total_clk: int,
    df1: pd.DataFrame | None = None,
    df2: pd.DataFrame | None = None,
    selected_language: str | None = None,
) -> tuple[list[dict], dict, dict]:
    """
    Build the App/URL sheet rows.

    The UI-selected language is the single source of truth. User URLs are
    validated only against DB rows matching the selected language. If a user
    URL is not present in DB under that language, it is auto-added to DB and
    then included as validated.
    """
    info: dict = {
        "warnings": [],
        "selected_language": selected_language,
        "validated_urls": [],
        "validated_urls_count": 0,
        "urls_added_to_db": [],
        "db_sheet_used": None,
        "pivot_urls_found": 0,
    }

    _JUNK = {"nan", "none", "", "app/url", "app", "url", "site", "domain",
             "language or line item"}
    user_apps: list[str] = []
    seen_user: set[str] = set()
    for raw in (app_urls_str or "").split("\n"):
        cu = _clean_url(raw)
        if cu and cu not in _JUNK and cu not in seen_user:
            user_apps.append(cu)
            seen_user.add(cu)

    # ── Resolve the single source-of-truth language ───────────────────────
    # 1) UI selection wins if provided.
    # 2) Otherwise try to auto-detect from File 1 (back-compat fallback).
    language: str | None = (selected_language or "").strip() or None
    if not language and df1 is not None and not df1.empty:
        language = detect_language_from_df1(df1)

    info["db_sheet_used"] = language or "all_sheets_fallback"

    # ── Step 2: validate user URLs against DB rows matching selected lang ──
    # Auto-add any user URL missing from that language section to the DB.
    db_records = _load_db_records()
    lang_norm  = _norm_lang(language) if language else None
    db_urls_for_lang: set[str] = {
        _clean_url(r["url"]) for r in db_records
        if lang_norm and _norm_lang(r["language"]) == lang_norm
    }

    validated_urls: list[str] = []
    added_to_db: list[str] = []
    validated_seen: set[str] = set()
    for u in user_apps:
        if lang_norm and u not in db_urls_for_lang:
            if add_url_to_db(u, language):
                added_to_db.append(u)
                db_urls_for_lang.add(u)
        if u not in validated_seen:
            validated_urls.append(u)
            validated_seen.add(u)

    info["validated_urls"]       = validated_urls
    info["validated_urls_count"] = len(validated_urls)
    info["urls_added_to_db"]     = added_to_db
    if added_to_db:
        info["warnings"].append(
            f"{len(added_to_db)} new URL(s) added to DB under "
            f"language='{language}'."
        )

    validated_set = set(validated_urls)

    # Pull the DB pool for the selected language (for fillers). If no language
    # provided, fall back to the legacy broad priority lookup.
    if language:
        pri_pool = [
            _clean_url(r["url"]) for r in db_records
            if _norm_lang(r["language"]) == lang_norm
            and _clean_url(r["url"]) not in _JUNK
            and _clean_url(r["url"]) not in validated_set
        ]
        all_pool = [
            _clean_url(r["url"]) for r in db_records
            if _clean_url(r["url"]) not in _JUNK
            and _clean_url(r["url"]) not in validated_set
        ]
        db_pools = {"primary": pri_pool, "broad": [], "multi": [], "all": all_pool}
    else:
        db_pools = get_db_urls_by_priority(language)

    def _filter(pool):
        return [u for u in pool if u not in _JUNK and u not in validated_set]

    pri_pool   = _filter(db_pools["primary"])
    broad_pool = _filter(db_pools["broad"])
    multi_pool = _filter(db_pools["multi"])
    fallback_p = _filter(db_pools["all"])

    random.shuffle(pri_pool)
    random.shuffle(broad_pool)
    random.shuffle(multi_pool)
    random.shuffle(fallback_p)

    pivot_urls: list[str] = []
    if df2 is not None and not df2.empty and df1 is not None and not df1.empty:
        pivot_urls = [
            _clean_url(u) for u in extract_pivot_urls(df1, df2)
            if _clean_url(u) not in _JUNK and _clean_url(u) not in validated_set
        ]
        random.shuffle(pivot_urls)
        info["pivot_urls_found"] = len(pivot_urls)

    filler_pool: list[str] = []
    seen_filler: set[str]  = set()
    for pool in [pivot_urls, pri_pool, broad_pool, multi_pool, fallback_p]:
        for u in pool:
            if u not in seen_filler and u not in validated_set:
                filler_pool.append(u)
                seen_filler.add(u)

    # ── Step 3: target URL count based on total impressions ───────────────
    if total_imp < 50_000:
        target_count = random.randint(30, 40)
    elif total_imp < 70_000:
        target_count = random.randint(50, 60)
    elif total_imp < 100_000:
        target_count = random.randint(60, 70)
    elif total_imp < 1_000_000:
        target_count = random.randint(70, 100)
    else:
        target_count = random.randint(110, 120)

    real_total = len(validated_urls) + len(filler_pool)
    if real_total > 0:
        target_count = min(target_count, real_total)

    n_val  = len(validated_urls)
    num_top = min(n_val, 15)

    top_val_urls   = validated_urls[:num_top]
    remaining_val  = validated_urls[num_top:]

    if num_top > 1:
        zone_size = 15 / num_top
        spread_indices: list[int] = []
        for i in range(num_top):
            z_start  = int(i * zone_size)
            z_end    = int((i + 1) * zone_size) - 1
            prev     = spread_indices[-1] if spread_indices else -1
            spread_indices.append(random.randint(max(z_start, prev + 1), z_end))
    elif num_top == 1:
        spread_indices = [random.randint(0, 5)]
    else:
        spread_indices = []

    top_block: list[str | None] = [None] * 15
    for idx, url in zip(spread_indices, top_val_urls):
        top_block[idx] = url

    fil_idx = 0
    for i in range(15):
        if top_block[i] is None:
            if fil_idx < len(filler_pool):
                top_block[i] = filler_pool[fil_idx]
                fil_idx += 1

    all_apps: list[str] = [u for u in top_block if u is not None]

    tail_pool = remaining_val + filler_pool[fil_idx:]
    needed    = max(0, target_count - len(all_apps))
    all_apps += tail_pool[:needed]

    if not all_apps:
        all_apps = user_apps[:]
        info["warnings"].append(
            "No validated or DB URLs found — falling back to raw user-supplied URLs."
        )

    n = len(all_apps)

    imp_weights: list[float] = []
    for i in range(n):
        if i < 8:
            weight = random.uniform(30.0, 50.0)
        elif i < 20:
            weight = random.uniform(12.0, 22.0)
        elif i < 85:
            weight = random.uniform(1.5, 6.0)
        else:
            weight = random.uniform(0.3, 1.2)
        imp_weights.append(weight)

    w_total = sum(imp_weights)
    if w_total > total_imp:
        scale = total_imp / w_total
        imp_weights = [w * scale for w in imp_weights]
        w_total = total_imp

    def largest_remainder(weights: list[float], w_tot: float,
                          total: int, min_val: int = 1) -> list[int]:
        if not weights: return []
        if w_tot == 0: w_tot = 1.0
        exact   = [(w / w_tot) * total for w in weights]
        floored = (
            [int(round(e)) for e in exact]
            if total < len(weights) * min_val
            else [max(min_val, int(round(e))) if w > 0 else 0
                  for w, e in zip(weights, exact)]
        )
        diff = total - sum(floored)
        if diff:
            fracs = [e - int(e) for e in exact]
            order = sorted(range(len(weights)),
                           key=lambda i: fracs[i], reverse=(diff > 0))
            for i in range(abs(diff)):
                idx = order[i % len(order)]
                if diff > 0:
                    floored[idx] += 1
                elif floored[idx] > min_val:
                    floored[idx] -= 1
        return floored

    app_imps = largest_remainder(imp_weights, w_total, total_imp, min_val=1)
    app_imps = deduplicate_preserving_sum(app_imps, gap=7)

    ranked_indices = sorted(range(n), key=lambda idx: app_imps[idx], reverse=True)

    raw_clks: list[float] = [0.0] * n
    for rank, idx in enumerate(ranked_indices, 1):
        imp = app_imps[idx]
        if imp < 180: continue
        if rank <= 8:
            raw_clks[idx] = imp * random.uniform(0.0045, 0.0052)
        elif rank <= 85:
            raw_clks[idx] = imp * random.uniform(0.0035, 0.0056)
        else:
            raw_clks[idx] = imp * random.uniform(0.0035, 0.0056)

    w_clk = sum(raw_clks) or float(n)
    app_clks = largest_remainder(raw_clks, w_clk, total_clk, min_val=0)

    for i, imp in enumerate(app_imps):
        if imp >= 180 and app_clks[i] < 1 and sum(app_clks) < total_clk:
            app_clks[i] = 1

    drift = total_clk - sum(app_clks)
    if drift:
        eligible = list(range(n)); random.shuffle(eligible)
        if drift > 0:
            for _ in range(drift):
                tgt = next(
                    (i for i in eligible
                     if app_imps[i] > 0 and (app_clks[i]+1)/app_imps[i] <= 0.0056),
                    max(eligible, key=lambda i: app_imps[i])
                )
                app_clks[tgt] += 1
        else:
            to_remove = abs(drift)
            for idx in eligible:
                if not to_remove: break
                while app_clks[idx] > 1 and to_remove:
                    if (app_clks[idx]-1) / app_imps[idx] >= 0.0035:
                        app_clks[idx] -= 1; to_remove -= 1
                    else: break
            if to_remove:
                for idx in eligible:
                    if not to_remove: break
                    while app_clks[idx] > 1 and to_remove:
                        app_clks[idx] -= 1; to_remove -= 1

    rows: list[dict] = [
        {"App/URL": url, "Impressions": imp, "Clicks": clk, "CTR": pct(clk, imp)}
        for url, imp, clk in zip(all_apps, app_imps, app_clks)
    ]
    rows.sort(key=lambda x: str(x["App/URL"]).strip().lower())
    rows.sort(key=lambda x: x["Impressions"], reverse=True)

    total_row: dict = {
        "App/URL": "Grand Total",
        "Impressions": total_imp,
        "Clicks": total_clk,
        "CTR": pct(total_clk, total_imp),
    }
    return rows, total_row, info


def process_file(job_id: str, filepath1: Path, filepath2: Path | None,
                 app_urls: str = "", selected_language: str | None = None):
    print(f"[DEBUG] process_file thread started for Job: {job_id}")
    try:
        jobs[job_id]["status"] = "processing"

        ext1 = filepath1.suffix.lower()
        df1  = pd.read_csv(filepath1) if ext1 == ".csv" else pd.read_excel(filepath1)

        if df1.empty:
            raise ValueError("File 1 contains no data.")

        total_imp = int(df1.get("Impressions", pd.Series([0])).apply(safe_float).sum())
        total_clk = int(df1.get("Clicks",      pd.Series([0])).apply(safe_float).sum())
        ctr_reach = pct(total_clk, total_imp)

        df2 = None
        if filepath2 is not None:
            ext2 = filepath2.suffix.lower()
            df2  = pd.read_csv(filepath2) if ext2 == ".csv" else pd.read_excel(filepath2)
            if df2.empty:
                df2 = None
                jobs[job_id]["warnings"] = jobs[job_id].get("warnings", []) + ["File 2 is empty - Sheet 8 skipped."]

        s1        = build_sheet1_reach(total_imp, total_clk)
        s2, s2t   = build_sheet2_date(df1, total_imp, total_clk, ctr_reach)
        s3, s3t   = build_sheet3_timeofday(total_imp, total_clk, ctr_reach)
        s4, s4t   = build_sheet4_age(total_imp, total_clk, ctr_reach)
        s5, s5t   = build_sheet5_gender(total_imp, total_clk, ctr_reach)
        s6, s6t   = build_sheet6_device(total_imp, total_clk, ctr_reach)
        s7, s7t   = build_sheet7_exchange(total_imp, total_clk, ctr_reach)

        city_info     = {}
        creative_info = {}
        s8 = s8t = s9 = s9t = None
        if df2 is not None:
            s8, s8t, city_info     = build_sheet8_city(df1, df2, total_imp, total_clk, ctr_reach)
            s9, s9t, creative_info = build_sheet9_creative(df1, df2, total_imp, total_clk)

        output_path = OUTPUT_DIR / f"report_{job_id}.xlsx"
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "REACH"
        h1 = ["Impressions","Clicks","Click Rate (CTR)","Reach","Frequency"]
        f1 = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
        write_sheet(ws1, h1, s1, formulas=f1, alignments="center")

        ws2 = wb.create_sheet("DATE")
        h2 = ["Date","Impressions","Clicks","Click Rate (CTR)","Viewable Impressions",
              "Measurable Impressions","Viewability","Sum of Starts (Video)",
              "Sum of Complete Views (Video)","VCR (Completion Rate)"]
        f2 = {
            "Click Rate (CTR)": "{Clicks}/{Impressions}",
            "Viewability": "{Viewable Impressions}/{Measurable Impressions}",
            "VCR (Completion Rate)": "{Sum of Complete Views (Video)}/{Sum of Starts (Video)}"
        }
        write_sheet(ws2, h2, s2, s2t, formulas=f2, alignments="center")

        ws10 = wb.create_sheet("APP URL")
        s10, s10t, app_info = build_sheet10_apps(
            app_urls, total_imp, total_clk, df1, df2,
            selected_language=selected_language,
        )
        h10 = ["App/URL", "Impressions", "Clicks", "CTR"]
        f10 = {"CTR": "{Clicks}/{Impressions}"}
        align10 = {"App/URL": "left", "Impressions": "right", "Clicks": "right", "CTR": "right"}
        write_sheet(ws10, h10, s10, s10t, formulas=f10, alignments=align10)
        jobs[job_id]["app_info"] = app_info

        ws3 = wb.create_sheet("Time of day")
        h3 = ["Time of Day","Impressions","Clicks","Click Rate (CTR)"]
        f_ctr = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
        align3 = {"Time of Day": "left", "Impressions": "right", "Clicks": "right", "Click Rate (CTR)": "right"}
        write_sheet(ws3, h3, s3, s3t, formulas=f_ctr, alignments=align3)

        ws7 = wb.create_sheet("EXCHANGE")
        h7 = ["Exchange","Impressions","Clicks","Click Rate (CTR)"]
        align7 = {"Exchange": "left", "Impressions": "right", "Clicks": "right", "Click Rate (CTR)": "right"}
        write_sheet(ws7, h7, s7, s7t, formulas=f_ctr, alignments=align7)

        ws6 = wb.create_sheet("DEVICE")
        h6 = ["Device Type","Impressions","Clicks","Click Rate (CTR)"]
        align6 = {"Device Type": "left", "Impressions": "right", "Clicks": "right", "Click Rate (CTR)": "right"}
        write_sheet(ws6, h6, s6, s6t, formulas=f_ctr, alignments=align6)

        if s9 is not None:
            ws9 = wb.create_sheet("CREATIVE")
            h9 = ["Creative","Impressions","Clicks","Click Rate (CTR)"]
            align9 = {"Creative": "left", "Impressions": "right", "Clicks": "right", "Click Rate (CTR)": "right"}
            write_sheet(ws9, h9, s9, s9t, formulas=f_ctr, alignments=align9)

        if s8 is not None:
            ws8 = wb.create_sheet("CITY")
            h8 = ["City","Impressions","Clicks","Click Rate (CTR)"]
            align8 = {"City": "left", "Impressions": "right", "Clicks": "right", "Click Rate (CTR)": "right"}
            write_sheet(ws8, h8, s8, s8t, formulas=f_ctr, alignments=align8)
            jobs[job_id]["city_info"] = city_info

        ws4 = wb.create_sheet("AGE")
        h4 = ["Age","Impressions","Clicks","Click Rate (CTR)"]
        align4 = {"Age": "left", "Impressions": "right", "Clicks": "right", "Click Rate (CTR)": "right"}
        write_sheet(ws4, h4, s4, s4t, formulas=f_ctr, alignments=align4)

        ws5 = wb.create_sheet("GENDER")
        h5 = ["Gender","Impressions","Clicks","Click Rate (CTR)"]
        align5 = {"Gender": "left", "Impressions": "right", "Clicks": "right", "Click Rate (CTR)": "right"}
        write_sheet(ws5, h5, s5, s5t, formulas=f_ctr, alignments=align5)

        wb.save(output_path)
        jobs[job_id]["status"] = "done"
        jobs[job_id]["output"] = str(output_path)

    except Exception as exc:
        jobs[job_id]["status"] = "error"
        jobs[job_id]["error"]  = str(exc)
    finally:
        for fp in [filepath1, filepath2]:
            if fp:
                try:
                    fp.unlink(missing_ok=True)
                except Exception:
                    pass


# -- Routes --------------------------------------------------------------------
@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response


@app.route("/upload", methods=["POST", "OPTIONS"])
def upload():
    print(f"\n[DEBUG] --- New upload request received ---")
    if request.method == "OPTIONS":
        return jsonify({}), 200

    if "file" not in request.files:
        return jsonify({"error": "No file in request"}), 400

    file1 = request.files["file"]
    if not file1.filename:
        return jsonify({"error": "No file selected"}), 400

    if Path(file1.filename).suffix.lower() not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Only .csv, .xlsx, .xls files allowed"}), 400

    job_id     = uuid.uuid4().hex
    save_path1 = UPLOAD_DIR / f"{job_id}_1_{secure_filename(file1.filename)}"
    file1.save(str(save_path1))

    save_path2 = None
    if "file2" in request.files and request.files["file2"].filename:
        file2 = request.files["file2"]
        if Path(file2.filename).suffix.lower() not in ALLOWED_EXTENSIONS:
            return jsonify({"error": "File 2: only .csv, .xlsx, .xls files allowed"}), 400
        save_path2 = UPLOAD_DIR / f"{job_id}_2_{secure_filename(file2.filename)}"
        file2.save(str(save_path2))

    jobs[job_id] = {"status": "queued", "output": None, "error": None, "city_info": None, "app_info": None}
    app_urls          = request.form.get("app_urls", "")
    selected_language = (request.form.get("language") or "").strip() or None

    threading.Thread(
        target=process_file,
        args=(job_id, save_path1, save_path2, app_urls, selected_language),
        daemon=True
    ).start()

    return jsonify({"job_id": job_id}), 202


# ── Language management (UI ↔ DB sync) ───────────────────────────────────────
@app.route("/languages", methods=["GET", "OPTIONS"])
def list_languages_route():
    if request.method == "OPTIONS":
        return jsonify({}), 200
    return jsonify({"languages": list_languages_for_ui()})


@app.route("/languages", methods=["POST", "OPTIONS"])
def add_language_route():
    if request.method == "OPTIONS":
        return jsonify({}), 200
    data = request.get_json(silent=True) or {}
    language = str(data.get("language", "")).strip()
    if not language:
        return jsonify({"error": "language required"}), 400
    added = add_language_to_db(language)
    return jsonify({
        "ok": True,
        "added_to_db": added,
        "languages": list_languages_for_ui(),
    })


@app.route("/languages/<path:language>", methods=["DELETE", "OPTIONS"])
def remove_language_route(language: str):
    if request.method == "OPTIONS":
        return jsonify({}), 200
    removed = remove_language_from_db(language)
    return jsonify({
        "ok": True,
        "removed_from_db": removed,
        "languages": list_languages_for_ui(),
    })


@app.route("/status/<job_id>")
def status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Unknown job"}), 404
    return jsonify({
        "status":    job["status"],
        "error":     job.get("error"),
        "city_info": job.get("city_info"),
        "app_info":  job.get("app_info"),
    })


@app.route("/debug/<job_id>")
def debug_job(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Unknown job"}), 404
    return jsonify({
        "status":    job["status"],
        "error":     job.get("error"),
        "city_info": job.get("city_info"),
        "app_info":  job.get("app_info"),
    })


@app.route("/download/<job_id>")
def download(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Unknown job"}), 404
    if job["status"] != "done":
        return jsonify({"error": "Report not ready"}), 400

    output_path = Path(job["output"])
    if not output_path.exists():
        return jsonify({"error": "Output file missing"}), 500

    @after_this_request
    def cleanup(response):
        def remove():
            time.sleep(60)
            try:
                output_path.unlink(missing_ok=True)
                del jobs[job_id]
            except Exception:
                pass
        threading.Thread(target=remove, daemon=True).start()
        return response

    return send_file(output_path, as_attachment=True,
                     download_name=f"campaign_report_{job_id[:8]}.xlsx")


# -- Entry point ---------------------------------------------------------------
if __name__ == "__main__":
    print("-  Ad Campaign Report Server running at http://localhost:5000")
    print("    Open Index.html in your browser to use the UI.")
    app.run(debug=True, port=5000)