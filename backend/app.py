# -*- coding: utf-8 -*-
"""
Ad Campaign Report Generator - Flask + openpyxl
Run: pip install flask openpyxl pandas werkzeug && python App.py
"""

import os
import re
import math
import uuid
import time
import random
import threading
import base64
from pathlib import Path
from datetime import datetime, timezone

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.errors import IgnoredError
from flask import Flask, request, jsonify, send_file, after_this_request
from flask_cors import CORS
from werkzeug.utils import secure_filename

# -- DB Config ----------------------------------------------------------------
_DB_LOCK = threading.Lock()
_DB_MASTER_SHEET  = "Sheet1"
_DB_URL_COL       = "URL / App Name"
_DB_LANG_COL      = "Language or Line Item"
_DB_ID_COL        = "ID"
_DB_HEADER_ROW    = 2
_DB_JUNK          = {"nan", "none", "", "app/url", "app", "url", "url / app name",
                     "language or line item"}

CURATED_LANGUAGES = [
    "Chinese", "Arabic", "Vietnamese", "Korean", "Japanese", "Turkish", "Spanish",
    "Singaporean", "Malay", "Sudanesh", "Indonesian", "Filipino OR Tagalog",
    "Thai", "Taiwan", "Hong kong","Hindi", "Italian", "French", "African", "Greek",
    "Mecedonians", "Croatians", "Nepalese", "(Dari) OR (Persian) OR (Farsi)",
    "Chin Haka", "Dinka", "Fiji", "Combodian OR Khmer", "Amharic",
    "(Burmash) OR (Karen)", "Assyrian", "South East Asian",
    "(African American) OR (African)", "Nigerians", "Ghanaians", "Egyptians",
    "Kenyans", "South Africa", "Moroccans", "Somalia", "Algerians", "Senegal",
    "Ethiopians", "Cameroons", "DR Congo", "Tanzania", "African other sites",
    "Pacific Islander", "New Zealand", "Samoa", "South Asian",
    "Indian-Tamil OR Tamil", "Indian Gujarati OT Gujarati",
    "Indian Malayalam OR Malayalam", "Indian Telugu - Telugu",
    "Indian_Punjabi OR Punjabi", "PUNJABI (CA) OR Punjabi Canada",
    "Pakistan", "Srilankan", "Bangaladesh", "Chinese_AU OR Chinese  Australia",
    "Chinesh_NZ OR Chinesh New Zealand", "Mandarin",
    "Mandarin CA OR Mandarin Canada", "Cantonese", "Cantonese Canada",
]

# -- Config -------------------------------------------------------------------
app = Flask(__name__)
CORS(app)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

BASE_DIR = Path(__file__).resolve().parent

UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
MAX_REPORT_AGE = 60 * 60 * 24  # keep generated outputs for 24 hours

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
jobs: dict[str, dict] = {}

APP_DB_FILE        = str(BASE_DIR / "data" / "App_Url Data base.xlsx")
CITY_REF_FILE      = str(BASE_DIR / "data" / "City for Aoutomation.xlsx")
CITY_REF_SHEET     = "Master Database"

# -- Excel styling helpers -----------------------------------------------------
HEADER_BG = "00B0F0"
HEADER_FG = "000000"
TOTAL_BG  = "9BC2E6"
TOTAL_FG  = "000000"

def _thin_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(ws, row, num_cols):
    fill = PatternFill("solid", fgColor=HEADER_BG)
    font = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 14.50

def style_data(ws, start_row, end_row, num_cols):
    font = Font(size=10, name="Calibri")
    for r in range(start_row, end_row):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.border = _thin_border()

def style_total_row(ws, row, num_cols):
    fill = PatternFill("solid", fgColor=TOTAL_BG)
    font = Font(bold=True, color=TOTAL_FG, size=11, name="Calibri")
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = _thin_border()

def auto_fit(ws):
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 50)

def write_sheet(ws, headers, rows, total_row=None, formulas=None, alignments=None, total_alignments=None):
    ws.sheet_view.showGridLines = True
    formulas = formulas or {}
    alignments = alignments or {}
    col_map = {h: get_column_letter(i) for i, h in enumerate(headers, 1)}

    def get_align_str(h_name, idx, align_config):
        if isinstance(align_config, str):
            return align_config
        if h_name in align_config:
            return align_config[h_name]
        if not align_config:
            return "left" if idx == 1 else "right"
        return "center"

    # Write headers
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        target = get_align_str(h, c_idx, alignments)
        cell.alignment = Alignment(horizontal=target, vertical="center")
    style_header(ws, 1, len(headers))

    def _assign(cell, val, h_name=None, row_idx=None):
        if hasattr(val, "item"):
            val = val.item()
        if h_name in formulas and row_idx is not None:
            f_temp = formulas[h_name]
            for h_ref, let in col_map.items():
                f_temp = f_temp.replace(f"{{{h_ref}}}", f"{let}{row_idx}")
            cell.value = f"={f_temp}"
            if h_name in ["CTR", "Click Rate (CTR)", "Viewability", "VCR (Completion Rate)"]:
                cell.number_format = "0.00%"
            return
        if isinstance(val, str) and val.strip().endswith("%"):
            try:
                cell.value = float(val.strip()[:-1]) / 100.0
                cell.number_format = "0.00%" if "." in val else "0%"
            except Exception:
                cell.value = val
        elif isinstance(val, str) and val.lstrip('-').isdigit():
            cell.value = int(val)
        else:
            cell.value = val

    # Write data rows
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            _assign(cell, row.get(h, ""), h_name=h, row_idx=r_idx)
            target = get_align_str(h, c_idx, alignments)
            cell.alignment = Alignment(horizontal=target, vertical="center")
            cell.border = _thin_border()

    style_data(ws, 2, 2 + len(rows), len(headers))

    # Write total row
    if total_row:
        t_row = 2 + len(rows)
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=t_row, column=c_idx)
            h_name = headers[c_idx - 1]
            val = total_row.get(h_name, "")
            
            # Use total_alignments if provided, else fall back to alignments
            t_align_config = total_alignments if total_alignments is not None else alignments
            target = get_align_str(h, c_idx, t_align_config)
            cell.alignment = Alignment(horizontal=target, vertical="center")

            if h_name in formulas:
                f_temp = formulas[h_name]
                for h_ref, let in col_map.items():
                    f_temp = f_temp.replace(f"{{{h_ref}}}", f"{let}{t_row}")
                cell.value = f"={f_temp}"
                cell.number_format = "0.00%"
            elif h_name in ["Impressions", "Clicks", "Reach", "Measurable Impressions",
                            "Viewable Impressions", "Sum of Starts (Video)",
                            "Sum of Complete Views (Video)"]:
                let = col_map[h_name]
                cell.value = f"=SUM({let}2:{let}{t_row-1})"
                cell.number_format = "#,##0"
            elif val is not None and val != "" and h_name != headers[0]:
                _assign(cell, val)
            else:
                _assign(cell, val)
        style_total_row(ws, t_row, len(headers))

    auto_fit(ws)

    try:
        ws.ignored_errors.append(IgnoredError(sqref="A1:ZZ5000",
                                              numberStoredAsText=True,
                                              evalError=True,
                                              emptyCellReference=True,
                                              calculatedColumn=True))
    except Exception:
        pass


# -- JS-Python utility functions -----------------------------------------------
def rand_float(mn, mx):
    return mn + random.random() * (mx - mn)

def rand_int(mn, mx):
    return random.randint(mn, mx)

def pct(num, den, decimals=2):
    if den == 0:
        return "0.00%"
    return f"{round((num / den) * 100, decimals):.{decimals}f}%"

def _largest_remainder(weights: list[float], total_weight: float, total_count: int) -> list[int]:
    """Distribute an integer total_count based on fractional weights."""
    if not weights or total_count <= 0:
        return [0] * len(weights)
    if total_weight <= 0:
        total_weight = sum(weights) or 1.0
    fractions = [(w / total_weight) * total_count for w in weights]
    integers = [int(f) for f in fractions]
    remainders = [(f - i, idx) for idx, (f, i) in enumerate(zip(fractions, integers))]
    remainders.sort(key=lambda x: x[0], reverse=True)
    gap = total_count - sum(integers)
    for i in range(int(gap)):
        integers[remainders[i][1]] += 1
    return integers

def _distribute_clicks(impressions: list[int], ctr_min: float = 0.0035, ctr_max: float = 0.0056) -> list[int]:
    """Distribute clicks based on impressions using random CTR within range."""
    return [round(imp * random.uniform(ctr_min, ctr_max)) for imp in impressions]

def safe_float(val, default=0.0):
    if val is None:
        return default
    s = str(val).strip().replace(',', '').replace(' ', '')
    if s.endswith('%'):
        try:
            return float(s[:-1]) / 100.0
        except ValueError:
            return default
    try:
        return float(s)
    except ValueError:
        return default

def safe_int(val, default=0):
    return int(safe_float(val, default))

def serial_to_date(serial):
    try:
        serial = float(serial)
        utc_days = int(serial - 25569)
        dt = datetime(1970, 1, 1, tzinfo=timezone.utc) + \
             __import__('datetime').timedelta(days=utc_days)
        return dt.strftime("%d %B, %Y")
    except Exception:
        return str(serial)


def deduplicate_preserving_sum(values: list[int], gap: int = 1) -> list[int]:
    _n = len(values)
    if _n < 2: return values
    items = [[i, v] for i, v in enumerate(values)]
    items.sort(key=lambda x: x[1], reverse=True)
    used: set[int] = set()

    for k in range(_n):
        val = items[k][1]
        if val not in used and all(abs(val - u) >= gap for u in used):
            used.add(val); continue
        found = False
        for _gap in sorted([gap, 5, 3, 2, 1], reverse=True):
            if _gap > gap: continue
            for _ in range(50):
                off  = random.randint(_gap, max(_gap + 5, _n))
                sign = random.choice([-1, 1])
                cand = val + sign * off
                if cand >= 1 and cand not in used and all(abs(cand - u) >= _gap for u in used):
                    items[k][1] = cand; used.add(cand); found = True; break
            if found: break
        if not found:
            off = 1
            while True:
                for cand in [val + off, val - off]:
                    if cand >= 1 and cand not in used:
                        items[k][1] = cand; used.add(cand); found = True; break
                if found: break
                off += 1

    drift = sum(values) - sum(x[1] for x in items)
    if drift:
        items.sort(key=lambda x: x[1], reverse=True)
        for k in range(_n):
            if not drift: break
            step = 1 if drift > 0 else -1
            cand = items[k][1] + step
            if cand >= 1 and cand not in used:
                items[k][1] = cand; used.add(cand); drift -= step
        if drift:
            items[0][1] += drift
    items.sort(key=lambda x: x[0])
    return [x[1] for x in items]

# -- URL cleaning utility (module-level) ---------------------------------------
def _clean_url(u: str) -> str:
    u = str(u).lower().strip()
    for prefix in ("https://", "http://", "www."):
        if u.startswith(prefix):
            u = u[len(prefix):]
    return u.rstrip("/")


# -- Line Item column auto-detection ------------------------------------------
LINE_ITEM_CANDIDATES = [
    "line item name",
    "line item",
    "lineitem",
    "line_item",
    "campaign line item",
    "ad name",
    "ad set name",
    "adset name",
    "placement",
    "ad unit",
]

_LI_EXCLUDE_KEYWORDS = {"id", "code", "key", "number", "num", "no"}

def detect_line_item_col(df: pd.DataFrame) -> str | None:
    cols_lower = {c.strip().lower(): c for c in df.columns}
    for candidate in LINE_ITEM_CANDIDATES:
        if candidate in cols_lower:
            return cols_lower[candidate]
    for lower, original in cols_lower.items():
        if "line" in lower and "item" in lower:
            tokens = set(re.split(r'[\s_\-]+', lower))
            if not tokens & _LI_EXCLUDE_KEYWORDS:
                return original
    return None


# -- Database helpers ----------------------------------------------------------
_DB_URL_COL  = "URL / App Name"
_DB_LANG_COL = "Language or Line Item"
_DB_JUNK     = {"nan", "none", "", "app/url", "app", "url", "url / app name"}


def _resolve_sheet_name(sheet_name: str) -> str | None:
    """Return the exact sheet name in the DB file, using case-insensitive match."""
    if not sheet_name or not sheet_name.strip():
        return None
    try:
        wb = openpyxl.load_workbook(APP_DB_FILE, read_only=True)
        names = wb.sheetnames
        wb.close()
        # Exact match first
        if sheet_name in names:
            return sheet_name
        # Case-insensitive fallback
        name_lower = sheet_name.strip().lower()
        for s in names:
            if s.strip().lower() == name_lower:
                return s
    except Exception:
        pass
    return None


def _load_db_sheet(sheet_name: str) -> pd.DataFrame:
    """Load a specific sheet from the DB file (header on row 2)."""
    resolved = _resolve_sheet_name(sheet_name)
    if not resolved:
        return pd.DataFrame()
    try:
        df = pd.read_excel(APP_DB_FILE, sheet_name=resolved, header=1)
        return df
    except Exception:
        return pd.DataFrame()


def get_db_sheet_names() -> list[str]:
    """Return all sheet names from the DB file for the UI dropdown."""
    try:
        wb = openpyxl.load_workbook(APP_DB_FILE, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


def get_urls_from_multiple_sheets(sheet_names_str: str) -> list[str]:
    """Return all cleaned URLs from multiple DB sheets, merged and deduplicated."""
    if not sheet_names_str: return []
    sheet_names = [s.strip() for s in sheet_names_str.split(",") if s.strip()]
    
    all_urls = set()
    for sn in sheet_names:
        df = _load_db_sheet(sn)
        if df.empty or _DB_URL_COL not in df.columns:
            continue
        for u in df[_DB_URL_COL].dropna().astype(str).tolist():
            cleaned = _clean_url(u)
            if cleaned and cleaned not in _DB_JUNK:
                all_urls.add(cleaned)
    return list(all_urls)


def append_urls_to_sheet(sheet_name: str, new_urls: list[str]) -> None:
    """
    Append new_urls to the selected sheet in the DB file
    under the URL / App Name column if they don't already exist.
    """
    if not new_urls:
        return
    try:
        resolved = _resolve_sheet_name(sheet_name)
        if not resolved:
            return
        wb = openpyxl.load_workbook(APP_DB_FILE)
        if resolved not in wb.sheetnames:
            wb.close()
            return
        ws = wb[resolved]

        # Find the URL column index (header is on row 2 per _load_db_sheet)
        url_col_idx = None
        for col in ws.iter_cols(min_row=2, max_row=2):
            for cell in col:
                if str(cell.value).strip() == _DB_URL_COL:
                    url_col_idx = cell.column
                    break
            if url_col_idx:
                break

        if url_col_idx is None:
            wb.close()
            return

        # Collect existing URLs
        existing = set()
        for row in ws.iter_rows(min_row=3, min_col=url_col_idx, max_col=url_col_idx):
            for cell in row:
                val = _clean_url(str(cell.value or ""))
                if val not in _DB_JUNK:
                    existing.add(val)

        # Append only truly new ones
        for url in new_urls:
            if url not in existing and url not in _DB_JUNK:
                ws.append({url_col_idx: url})
                existing.add(url)

        wb.save(APP_DB_FILE)
        wb.close()
    except Exception as e:
        print(f"[WARN] append_urls_to_sheet failed: {e}")


def validate_and_prepare_urls(
    user_urls: list[str],
    sheet_names_str: str,
) -> tuple[list[str], list[str]]:
    """
    Check each user URL against all selected DB sheets.
    - Matched  → use as-is
    - Not found → append to the FIRST sheet in the selection, still use it
    """
    _JUNK = {"nan", "none", "", "app/url", "app", "url", "site", "domain"}
    cleaned_user = [u for u in (_clean_url(u) for u in user_urls) if u not in _JUNK]

    if not cleaned_user:
        return [], []

    # Get combined set from all selected sheets
    db_urls_set = set(get_urls_from_multiple_sheets(sheet_names_str))

    not_found = [u for u in cleaned_user if u not in db_urls_set]

    # Append new URLs to the FIRST selected sheet so they persist somewhere
    if not_found and sheet_names_str:
        primary_sheet = sheet_names_str.split(",")[0].strip()
        append_urls_to_sheet(primary_sheet, not_found)

    # All user URLs are valid to use
    all_to_use = cleaned_user  # preserves original order
    return all_to_use, not_found


# -- Language detection (kept for fallback reference, not used in Sheet 10) ----
LANGUAGE_KEYWORDS = {
    "cantonese":  "Cantonese",
    "mandarin":   "Mandarin",
    "chinese":    "Chinese",
    "hong kong":  "Cantonese",
    "hk":         "Cantonese",
    "taiwan":     "Mandarin",
    "tw":         "Mandarin",
    "arabic":     "Arabic",
    "english":    "English",
    "french":     "French",
    "spanish":    "Spanish",
    "hindi":      "Hindi",
    "malay":      "Malay",
    "tamil":      "Tamil",
    "tagalog":    "Filipino/Tagalog",
    "filipino":   "Filipino/Tagalog",
    "indonesian": "Indonesian",
    "thai":       "Thai",
    "vietnamese": "Vietnamese",
    "korean":     "Korean",
    "japanese":   "Japanese",
    "italian":    "Italian",
    "urdu":       "Urdu",
    "bengali":    "Bengali",
    "persian":    "Dari/Persian",
    "dari":       "Dari/Persian",
    "turkish":    "Turkish",
    "greek":      "Greek",
    "punjabi":    "Punjabi",
    "gujarati":   "Gujarati",
    "malayalam":  "Malayalam",
    "telugu":     "Telugu",
    "sinhala":    "Sinhala/Tamil",
    "khmer":      "Khmer",
    "nepali":     "Nepali",
    "amharic":    "Amharic",
    "burmese":    "Burmese/Karen",
    "karen":      "Burmese/Karen",
    "fijian":     "Fijian",
    "dinka":      "Dinka",
    "croatian":   "Croatian",
    "macedonian": "Macedonian",
    "assyrian":   "Assyrian/Arabic",
    "portuguese": "Spanish",
    "german":     "Multi-language",
    "russian":    "Multi-language",
}

_DB_MASTER_SHEET = "\U0001f4cb Master Database"


def _normalize_text(val: str) -> str:
    val = str(val).strip()
    val = re.sub(r'[\u2013\u2014\u2212\u2015]', '-', val)
    val = re.sub(r'\s+', ' ', val)
    return val.lower().strip()


def _clean_li2(val: str) -> str:
    val = str(val).strip()
    val = re.sub(r'^[\w\-]+\s*\|\s*', '', val)
    return _normalize_text(val)


def _normalize_city(val: str) -> str:
    return re.sub(r'\s+', ' ', str(val).strip()).title()


# -- Sheet builders ------------------------------------------------------------

def build_sheet1_reach(total_imp, total_clk):
    ctr = pct(total_clk, total_imp)
    freq = 3
    reach = int(total_imp / freq) + random.randint(200, 300)
    return [{
        "Impressions": total_imp,
        "Clicks": total_clk,
        "Click Rate (CTR)": ctr,
        "Reach": reach,
        "Frequency": freq
    }]


def build_sheet2_date(df, total_imp, total_clk, ctr_reach, is_banner=False):
    rows = []
    sum_imp = sum_clk = sum_view = sum_meas = sum_starts = sum_comp = 0
    vcr_weighted = vcr_imp_total = 0

    for _, r in df.iterrows():
        imp   = safe_int(r.get("Impressions", 0))
        clk   = safe_int(r.get("Clicks", 0))

        date_val = r.get("Date", "")
        try:
            float(date_val)
            date_str = serial_to_date(date_val)
        except (ValueError, TypeError):
            date_str = str(date_val)

        sum_imp += imp
        sum_clk += clk

        if is_banner:
            # Generate viewability synthetically for banner format
            view = round(imp * random.uniform(0.55, 0.70))
            meas = round(imp * random.uniform(0.90, 0.95))
            sum_view += view
            sum_meas += meas
            rows.append({
                "Date": date_str,
                "Impressions": imp,
                "Clicks": clk,
                "Click Rate (CTR)": pct(clk, imp),
                "Viewable Impressions": view,
                "Measurable Impressions": meas,
                "Viewability": pct(view, meas),
            })
        else:
            view    = safe_int(r.get("Viewable Impressions", 0))
            meas    = safe_int(r.get("Measurable Impressions", 0))
            starts  = safe_int(r.get("Start views", 0))
            comps   = safe_int(r.get("Complete Views", 0))
            vcr_raw = safe_float(r.get("Video Completion Rate (VCR)", 0))

            sum_view   += view;  sum_meas   += meas
            sum_starts += starts; sum_comp  += comps
            vcr_weighted  += vcr_raw * imp
            vcr_imp_total += imp

            vcr_pct = f"{round(vcr_raw * 100)}%" if vcr_raw > 0 else "0%"

            rows.append({
                "Date": date_str,
                "Impressions": imp,
                "Clicks": clk,
                "Click Rate (CTR)": pct(clk, imp),
                "Viewable Impressions": view,
                "Measurable Impressions": meas,
                "Viewability": pct(view, meas),
                "Sum of Starts (Video)": starts,
                "Sum of Complete Views (Video)": comps,
                "VCR (Completion Rate)": vcr_pct
            })

    if is_banner:
        total = {
            "Date": "Grand Total",
            "Impressions": sum_imp,
            "Clicks": sum_clk,
            "Click Rate (CTR)": ctr_reach,
            "Viewable Impressions": sum_view,
            "Measurable Impressions": sum_meas,
            "Viewability": pct(sum_view, sum_meas),
        }
    else:
        avg_vcr = f"{round((vcr_weighted / vcr_imp_total) * 100)}%" if vcr_imp_total > 0 else "0%"
        total = {
            "Date": "Grand Total",
            "Impressions": sum_imp,
            "Clicks": sum_clk,
            "Click Rate (CTR)": ctr_reach,
            "Viewable Impressions": sum_view,
            "Measurable Impressions": sum_meas,
            "Viewability": pct(sum_view, sum_meas),
            "Sum of Starts (Video)": sum_starts,
            "Sum of Complete Views (Video)": sum_comp,
            "VCR (Completion Rate)": avg_vcr
        }

    return rows, total


def build_sheet3_timeofday(total_imp, total_clk, ctr_reach):
    hour_weights = {
        0:0.5,1:0.4,2:0.3,3:0.3,4:0.4,5:0.7,6:1.2,7:1.5,
        8:2.0,9:2.5,10:2.3,11:2.4,12:2.2,13:1.8,14:1.6,15:1.7,
        16:1.9,17:2.6,18:2.8,19:2.5,20:2.0,21:1.6,22:1.2,23:0.8
    }

    noisy = [hour_weights[h] * (0.85 + random.random() * 0.30) for h in range(24)]
    total_w = sum(noisy)
    hourly_imp = [int((w / total_w) * total_imp) for w in noisy]
    diff = total_imp - sum(hourly_imp)
    for _ in range(abs(diff)):
        hourly_imp[random.randint(0, 23)] += 1 if diff > 0 else -1

    hourly_imp = deduplicate_preserving_sum(hourly_imp, gap=1)

    hourly_clk = []
    for imp in hourly_imp:
        if imp >= 180:
            c_min = math.ceil(imp * 0.0035)
            c_max = math.floor(imp * 0.0056)
            if c_min > c_max:
                hourly_clk.append(c_min)
            else:
                hourly_clk.append(random.randint(c_min, c_max))
        else:
            hourly_clk.append(0)

    current = sum(hourly_clk)
    diff_clk = total_clk - current
    eligible = [i for i, imp in enumerate(hourly_imp) if imp >= 180]

    iterations = 0
    while diff_clk != 0 and eligible and iterations < 5000:
        iterations += 1
        idx = random.choice(eligible)
        imp = hourly_imp[idx]
        clk = hourly_clk[idx]
        if diff_clk > 0:
            if (clk + 1) / imp <= 0.0056:
                hourly_clk[idx] += 1; diff_clk -= 1
        elif diff_clk < 0:
            if clk > 1 and (clk - 1) / imp >= 0.0035:
                hourly_clk[idx] -= 1; diff_clk += 1

    if diff_clk != 0 and eligible:
        for _ in range(abs(diff_clk)):
            safe_indices = []
            for i in eligible:
                if diff_clk > 0 and (hourly_clk[i] + 1) / hourly_imp[i] <= 0.0056:
                    safe_indices.append(i)
                elif diff_clk < 0 and hourly_clk[i] > 1 and (hourly_clk[i] - 1) / hourly_imp[i] >= 0.0035:
                    safe_indices.append(i)
            if not safe_indices: break
            idx = random.choice(safe_indices)
            if diff_clk > 0:
                hourly_clk[idx] += 1; diff_clk -= 1
            else:
                hourly_clk[idx] -= 1; diff_clk += 1

    rows = []
    for h in range(24):
        imp = hourly_imp[h]; clk = hourly_clk[h]
        rows.append({
            "Time of Day": h,
            "Impressions": imp,
            "Clicks": clk,
            "Click Rate (CTR)": pct(clk, imp)
        })

    total = {
        "Time of Day": "Grand Total",
        "Impressions": total_imp,
        "Clicks": total_clk,
        "Click Rate (CTR)": ctr_reach
    }
    return rows, total


def build_sheet4_age(total_imp, total_clk, ctr_reach):
    pct_1824 = rand_float(0.20, 0.25)
    pct_4554 = rand_float(0.09, 0.12)
    remaining = max(0.60, min(0.74, 1 - pct_1824 - pct_4554))
    split = rand_float(0.45, 0.55)
    pct_2534 = remaining * split
    pct_3544 = remaining - pct_2534

    imp_1824 = round(total_imp * pct_1824)
    imp_4554 = round(total_imp * pct_4554)
    rem_imp  = total_imp - imp_1824 - imp_4554
    imp_2534 = round(rem_imp * split)
    imp_3544 = rem_imp - imp_2534
    imp_2534 += total_imp - (imp_1824 + imp_2534 + imp_3544 + imp_4554)

    groups = [
        {"age": "18-24", "imp": imp_1824},
        {"age": "25-34", "imp": imp_2534},
        {"age": "35-44", "imp": imp_3544},
        {"age": "45-54", "imp": imp_4554},
    ]

    clicks = [round(g["imp"] * rand_float(0.0035, 0.0056)) if g["imp"] >= 180 else 0 for g in groups]
    c_sum = sum(clicks)
    if c_sum > 0:
        clicks = [round(c * total_clk / c_sum) if groups[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]

    drift = total_clk - sum(clicks)
    clicks[clicks.index(max(clicks))] += drift

    rows = [{
        "Age": g["age"],
        "Impressions": g["imp"],
        "Clicks": clicks[i],
        "Click Rate (CTR)": pct(clicks[i], g["imp"])
    } for i, g in enumerate(groups)]

    total = {"Age": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


def build_sheet5_gender(total_imp, total_clk, ctr_reach):
    female_pct = rand_float(0.30, 0.40)
    imp_female = round(total_imp * female_pct)
    imp_male   = total_imp - imp_female
    male_pct   = imp_male / total_imp

    if male_pct < 0.58:
        imp_male   = round(total_imp * rand_float(0.58, 0.61))
        imp_female = total_imp - imp_male
    elif male_pct > 0.65:
        imp_male   = round(total_imp * rand_float(0.62, 0.65))
        imp_female = total_imp - imp_male

    imp_male += total_imp - (imp_male + imp_female)

    groups = [{"gender": "Male", "imp": imp_male}, {"gender": "Female", "imp": imp_female}]
    clicks = [round(g["imp"] * rand_float(0.0035, 0.0056)) if g["imp"] >= 180 else 0 for g in groups]
    g_sum  = sum(clicks)
    if g_sum > 0:
        clicks = [round(c * total_clk / g_sum) if groups[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]

    drift = total_clk - sum(clicks)
    clicks[clicks.index(max(clicks))] += drift

    rows = [{
        "Gender": g["gender"],
        "Impressions": g["imp"],
        "Clicks": clicks[i],
        "Click Rate (CTR)": pct(clicks[i], g["imp"])
    } for i, g in enumerate(groups)]

    total = {"Gender": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


def build_sheet6_device(total_imp, total_clk, ctr_reach):
    tablet_pct  = rand_float(0.07, 0.10)
    desktop_shr = rand_float(0.45, 0.55)
    mobile_shr  = 1 - desktop_shr
    rem         = 1 - tablet_pct

    imp_tablet  = round(total_imp * tablet_pct)
    imp_desktop = round(total_imp * rem * desktop_shr)
    imp_mobile  = total_imp - imp_tablet - imp_desktop
    imp_mobile += total_imp - (imp_desktop + imp_mobile + imp_tablet)

    groups = [
        {"device": "Desktop",     "imp": imp_desktop},
        {"device": "Smart Phone", "imp": imp_mobile},
        {"device": "Tablet",      "imp": imp_tablet},
    ]

    clicks = [max(0, int(g["imp"] * rand_float(0.0035, 0.0056) + rand_float(-0.4, 0.4)))
              if g["imp"] >= 180 else 0 for g in groups]
    d_sum  = sum(clicks)
    if d_sum > 0:
        clicks = [int(c * total_clk / d_sum) if groups[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]

    drift = total_clk - sum(clicks)
    clicks[clicks.index(max(clicks))] += drift

    rows = [{
        "Device Type": g["device"],
        "Impressions": g["imp"],
        "Clicks": clicks[i],
        "Click Rate (CTR)": pct(clicks[i], g["imp"])
    } for i, g in enumerate(groups)]

    total = {"Device Type": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


def build_sheet7_exchange(total_imp, total_clk, ctr_reach):
    tier1_other = ["BidSwitch","Index Exchange","Magnite DV+","PubMatic"]
    tier2 = ["Criteo Commerce Grid","Equativ","InMobi","Media.net","Nexxen (fka Unruly)","OpenX","Sovrn"]
    tier3 = ["Microsoft Monetize","Epsilon Core Private Exchange","Yieldmo","TripleLift",
             "Kargo","Improve Digital","TeadsTv","GumGum","Adform"]

    ex_t1_imp = int(total_imp * rand_float(0.80, 0.85))
    ex_t2_imp = int(total_imp * rand_float(0.08, 0.13))
    ex_t3_imp = total_imp - ex_t1_imp - ex_t2_imp

    google_imp   = int(ex_t1_imp * rand_float(0.35, 0.45))
    t1_other_imp = ex_t1_imp - google_imp

    def split_across(names, total):
        remaining = total
        result = []
        for i, name in enumerate(names):
            if i == len(names) - 1:
                imp = remaining
            else:
                imp = max(1, rand_int(1, int((remaining - (len(names) - 1 - i)) * 0.6)))
            remaining -= imp
            result.append({"name": name, "imp": imp})
        return result

    exchanges = (
        [{"name": "Google Ad Manager", "imp": google_imp}]
        + split_across(tier1_other, t1_other_imp)
        + split_across(tier2, ex_t2_imp)
        + split_across(tier3, ex_t3_imp)
    )

    drift = total_imp - sum(e["imp"] for e in exchanges)
    exchanges[0]["imp"] += drift

    clicks = [round(e["imp"] * rand_float(0.0035, 0.0056)) if e["imp"] >= 180 else 0 for e in exchanges]
    e_sum  = sum(clicks)
    if e_sum > 0:
        clicks = [round(c * total_clk / e_sum) if exchanges[i]["imp"] >= 180 else 0 for i, c in enumerate(clicks)]

    e_drift = total_clk - sum(clicks)
    max_idx = max(range(len(exchanges)), key=lambda i: exchanges[i]["imp"])
    clicks[max_idx] += e_drift

    rows = sorted([{
        "Exchange": e["name"],
        "Impressions": e["imp"],
        "Clicks": clicks[i],
        "Click Rate (CTR)": pct(clicks[i], e["imp"])
    } for i, e in enumerate(exchanges)], key=lambda x: x["Exchange"])

    total = {"Exchange": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}
    return rows, total


# -- City DB helpers -----------------------------------------------------------

def get_city_db_sheet_names() -> list[str]:
    """Return sheet names from City for Automation.xlsx, excluding summary sheets."""
    EXCLUDE = {"summary by state"}
    try:
        xl = pd.ExcelFile(CITY_REF_FILE)
        return [s for s in xl.sheet_names if s.strip().lower() not in EXCLUDE]
    except Exception as e:
        print(f"[WARN] get_city_db_sheet_names failed: {e}")
        return []


def load_city_db_sheet(sheet_names_str: str) -> list[dict]:
    """
    Load and merge (city_name, weight, creatives) from multiple sheets.
    Returns a list of dicts: {'name': str, 'weight': float, 'creatives': list[str]}
    """
    if not sheet_names_str: return []
    sheet_names = [sn.strip() for sn in sheet_names_str.split(",") if sn.strip()]
    
    seen_cities = {} # city_lower -> {'name': str, 'weight': float, 'creatives': set()}
    
    try:
        xl = pd.ExcelFile(CITY_REF_FILE)
        for sn in sheet_names:
            if sn not in xl.sheet_names:
                continue
            df = pd.read_excel(xl, sheet_name=sn)
            if "City" not in df.columns:
                continue
            
            weight_col = next(
                (c for c in ["Potential Impressions", "Unique Cookies w/ Impressions"]
                 if c in df.columns), None
            )
            creative_col = next(
                (c for c in df.columns if "creative" in c.lower()), None
            )
            
            df = df[df["City"].notna()].copy()
            for _, row in df.iterrows():
                city_name = str(row["City"]).strip()
                if not city_name or city_name.lower() in ["nan", "none", ""]:
                    continue
                
                weight = safe_float(row[weight_col]) if weight_col else 1.0
                creatives = []
                if creative_col and pd.notna(row[creative_col]):
                    # Handle multiple creatives in one cell if comma-separated
                    raw_c = str(row[creative_col])
                    creatives = [c.strip() for c in raw_c.replace("|", ",").split(",") if c.strip()]
                
                cl = city_name.lower()
                if cl in seen_cities:
                    seen_cities[cl]['weight'] += weight
                    seen_cities[cl]['creatives'].update(creatives)
                else:
                    seen_cities[cl] = {
                        'name': city_name,
                        'weight': weight,
                        'creatives': set(creatives)
                    }
        
        merged = [
            {
                'name': v['name'],
                'weight': float(v['weight']),
                'creatives': sorted(list(v['creatives']))
            } 
            for v in seen_cities.values()
        ]
        return sorted(merged, key=lambda x: x['weight'], reverse=True)
    except Exception as e:
        print(f"[WARN] load_city_db_sheet failed: {e}")
        return []


# -- City reference file helper -----------------------------------------------

def _load_city_reference(n_cities: int = 50) -> list[tuple[str, float]]:
    """
    Load top N cities and their weights from City for Aoutomation.xlsx.
    Uses 'Potential Impressions' as weight. Returns [(city_name, weight), ...].
    """
    try:
        df = pd.read_excel(CITY_REF_FILE, sheet_name=CITY_REF_SHEET)
        if "City" not in df.columns:
            return []
        weight_col = next(
            (c for c in ["Potential Impressions", "Unique Cookies w/ Impressions"]
             if c in df.columns), None
        )
        df = df[df["City"].notna()].copy()
        df["_city"] = df["City"].astype(str).str.strip()
        df = df[df["_city"].str.lower().notnull() & (df["_city"] != "")]
        if weight_col:
            df["_weight"] = df[weight_col].apply(safe_float)
        else:
            df["_weight"] = 1.0
        df = df[df["_weight"] > 0].sort_values("_weight", ascending=False)
        return [(row["_city"], float(row["_weight"]))
                for _, row in df.head(n_cities).iterrows()]
    except Exception as e:
        print(f"[WARN] _load_city_reference failed: {e}")
        return []


# -- Sheet 8 - City -----------------------------------------------------------

def build_sheet8_city(df1: pd.DataFrame, df2: pd.DataFrame,
                      total_imp: int, total_clk: int, ctr_reach: str,
                      city_db_sheet: str = "Master Database",
                      is_banner: bool = False):
    info = {
        "matched_line_items": 0,
        "cities_found": 0,
        "city_source": "",
        "warnings": [],
        "debug": {}
    }

    # ── Shared distribution helpers ───────────────────────────────────────────
    def _largest_remainder(weights, w_total, total):
        if not weights or w_total == 0: return []
        exact   = [(w / w_total) * total for w in weights]
        floored = (
            [int(e) for e in exact]
            if total < len(weights)
            else [max(1, int(e)) if w > 0 else 0 for w, e in zip(weights, exact)]
        )
        remainder = total - sum(floored)
        fracs = [e - int(e) for e in exact]
        order = sorted(range(len(weights)), key=lambda i: fracs[i], reverse=(remainder > 0))
        for i in range(abs(remainder)):
            idx = order[i % len(order)]
            if remainder > 0:
                floored[idx] += 1
            elif floored[idx] > 1:
                floored[idx] -= 1
        return floored

    def _distribute_clicks(city_imps):
        """Distribute total_clk across cities with CTR strictly 0.35%–0.56%."""
        CTR_LO, CTR_HI = 0.0035, 0.0056
        raw = [max(1, int(imp * random.uniform(CTR_LO, CTR_HI))) if imp >= 180 else 0
               for imp in city_imps]
        raw_sum = sum(raw)
        if raw_sum == 0:
            return [0] * len(city_imps)
        clks = [max(1, int(round(r * total_clk / raw_sum))) if city_imps[i] >= 180 else 0
                for i, r in enumerate(raw)]
        drift = total_clk - sum(clks)
        if drift:
            eligible = [i for i, imp in enumerate(city_imps) if imp >= 180] or list(range(len(city_imps)))
            if drift > 0:
                clks[max(eligible, key=lambda i: city_imps[i])] += drift
            else:
                for _ in range(abs(drift)):
                    valid = [i for i in eligible if clks[i] > 1]
                    if not valid: break
                    clks[max(valid, key=lambda i: clks[i])] -= 1
        return clks

    def _build_rows(city_names, city_imps, city_clks):
        rows = [
            {"City": city, "Impressions": imp, "Clicks": clk,
             "Click Rate (CTR)": pct(clk, imp)}
            for city, imp, clk in zip(city_names, city_imps, city_clks) if imp > 0
        ]
        rows.sort(key=lambda x: str(x["City"]).strip().lower())
        return rows

    grand_total = {"City": "Grand Total", "Impressions": total_imp,
                   "Clicks": total_clk, "Click Rate (CTR)": ctr_reach}

    # ── Phase 1: Line Item Matching ───────────────────────────────────────────
    li_col1 = detect_line_item_col(df1)
    li_col2 = detect_line_item_col(df2)

    info["debug"]["li_col1"] = li_col1
    info["debug"]["li_col2"] = li_col2

    matched_df2 = pd.DataFrame()
    if li_col1 and li_col2:
        items1 = [str(v).strip() for v in df1[li_col1].dropna().unique() if str(v).strip()]

        def get_tokens(s):
            s = str(s).lower()
            # If format is "ID|Name", prioritize tokens from the Name part
            if "|" in s:
                s = s.split("|")[-1]
            return set(re.findall(r'[a-z0-9]+', s))

        def matches_any_line_item(v2_raw):
            v2_tokens = get_tokens(str(v2_raw))
            if not v2_tokens: return False
            for i1 in items1:
                i1_tokens = get_tokens(i1)
                if not i1_tokens: continue
                # Match if tokens overlap significantly or one is a subset
                if i1_tokens.issubset(v2_tokens) or v2_tokens.issubset(i1_tokens):
                    return True
            return False

        mask = df2[li_col2].apply(matches_any_line_item)
        matched_df2 = df2[mask].copy()
        info["matched_line_items"] = int(mask.sum())
        if matched_df2.empty:
            info["warnings"].append(
                f"Line item matching found 0 rows. Falling back to all File 2 rows."
            )

    if matched_df2.empty:
        matched_df2 = df2.copy()
        info["debug"]["fallback_used"] = True

def build_sheet8_city(matched_df1: pd.DataFrame, matched_df2: pd.DataFrame, 
                      total_imp: int, city_db_sheet: str, is_banner: bool = False):
    """
    Sheet 8: Distribution by City.
    Logic:
    1. Identify all cities from Input File 2.
    2. Extract unique creatives associated with each city from File 2.
    3. Look up related creatives from City DB Sheet.
    4. Use weights from Input File 2.
    5. Handle duplication to ensure valid distribution.
    """
    info = {"warnings": [], "debug": {}, "city_source": ""}
    rows = []

    # Detect necessary columns in File 1
    creative_col1 = next((c for c in matched_df1.columns if "creative" in c.lower() and "id" not in c.lower()), matched_df1.columns[0])
    
    # Detect necessary columns in File 2
    city_col = None
    for col in matched_df2.columns:
        if col.strip().lower() == "city": city_col = col; break
    if not city_col:
        for col in matched_df2.columns:
            if "city" in col.strip().lower(): city_col = col; break
    
    weight_col2 = next((c for c in matched_df2.columns if "impression" in c.lower() or "weight" in c.lower()), None)
    li_col2 = next((c for c in matched_df2.columns if "line item" in c.lower()), matched_df2.columns[0])

    if city_col is None:
        info["warnings"].append("No 'City' column detected in File 2.")
        return [], None, info

    # ── Phase 3: Load City DB Sheet ───────────────────────────────────────────
    db_cities_raw = load_city_db_sheet(city_db_sheet)
    if not db_cities_raw:
        db_cities_raw = load_city_db_sheet("Master Database")
        if db_cities_raw:
            info["warnings"].append(f"Sheet '{city_db_sheet}' not found. Using Master Database.")

    db_lookup = {c['name'].strip().lower(): c for c in db_cities_raw}

    # ── Phase 2: Filter File 2 by Strict Matching (Steps 1-3) ────────────────
    # Step 1: Extract unique Line Items from File 1
    li_col1 = next((c for c in matched_df1.columns if "line item" in c.lower()), matched_df1.columns[0])
    target_li_names = set(matched_df1[li_col1].dropna().astype(str).str.strip().unique())
    
    # Step 2: Normalize File 2 Line Items (Strip before |)
    # Step 3: Filter rows that match Step 1 exactly
    def normalize_and_match(val):
        val_str = str(val).strip()
        if "|" in val_str:
            val_str = val_str.split("|", 1)[1].strip()
        return val_str in target_li_names

    mask = matched_df2[li_col2].apply(normalize_and_match)
    filtered_df2 = matched_df2[mask].copy()
    
    # Debug info for the user
    info["debug"]["target_line_items"] = list(target_li_names)
    info["debug"]["filtered_rows"] = len(filtered_df2)

    if filtered_df2.empty:
        info["warnings"].append("No exact Line Item matches found after normalization. Using all data as fallback.")
        filtered_df2 = matched_df2.copy()

    # ── Phase 4: Extract Unique Cities (Step 4) ───────────────────────────────
    # Group by city to sum weights and collect unique creatives
    city_groups = filtered_df2.groupby(city_col).agg({
        weight_col2: 'sum',
        li_col2: lambda x: sorted(list(set(x.dropna().astype(str))))
    }).reset_index()

    input_cities_map = {}
    for _, row in city_groups.iterrows():
        name = str(row[city_col]).strip()
        if name and name.lower() not in ["nan", "none", ""]:
            input_cities_map[name.lower()] = {
                "name": name,
                "weight": safe_float(row[weight_col2]),
                "creatives": row[li_col2]
            }

    final_cities = []
    # Filter against selected DB sheet(s)
    for db_city in db_cities_raw:
        cl = db_city['name'].lower()
        if cl in input_cities_map:
            inp = input_cities_map[cl]
            final_cities.append({
                "name": inp['name'],
                "weight": inp['weight'],
                "creatives": db_city['creatives'] if db_city['creatives'] else inp['creatives']
            })

    # ── Phase 5: Smart Capping & Supplementation (Banner Only) ────────────────
    if is_banner:
        # Cap at 60 cities to keep the report professional
        if len(final_cities) > 60:
            final_cities.sort(key=lambda x: x["weight"], reverse=True)
            final_cities = final_cities[:60]
            
        # Supplement to reach at least 40 cities if needed
        if len(final_cities) < 40:
            existing_names = {c["name"].lower() for c in final_cities}
            master_cities = load_city_db_sheet("Master Database")
            added = 0
            for mc in master_cities:
                if added >= (40 - len(final_cities)): break
                if mc["name"].lower() not in existing_names:
                    final_cities.append({
                        "name": mc["name"],
                        "weight": mc["weight"],
                        "creatives": mc["creatives"] if mc["creatives"] else ["General Creative"]
                    })
                    added += 1

    # Final sort for report presentation
    final_cities.sort(key=lambda x: x["weight"], reverse=True)

    # ── Phase 6: Build Output Rows with Duplication ───────────────────────────
    all_f1_creatives = sorted(list(set(matched_df1[creative_col1].dropna().astype(str))))
    if not all_f1_creatives: all_f1_creatives = ["Default Creative"]

    for city_info in final_cities:
        city_name = city_info["name"]
        city_weight = max(city_info["weight"], 1.0)
        city_creatives = city_info["creatives"]
        
        if not city_creatives:
            city_creatives = all_f1_creatives
            
        weight_per_cr = city_weight / len(city_creatives)
        for cr in city_creatives:
            rows.append({
                "City": city_name,
                "Creative": cr,
                "_weight": weight_per_cr
            })

    if not rows:
        # Absolute fallback if filtering left us with nothing
        for idx, city_name in enumerate(["Sydney", "Melbourne", "Brisbane", "Perth"]):
            rows.append({"City": city_name, "Creative": all_f1_creatives[0], "_weight": 1.0})

    # Convert to DF and scale
    df_rows = pd.DataFrame(rows)
    total_weight = df_rows["_weight"].sum()
    scaled_imps = _largest_remainder(df_rows["_weight"].tolist(), total_weight, total_imp)
    scaled_imps = deduplicate_preserving_sum(scaled_imps, gap=1)
    
    df_rows["Impressions"] = scaled_imps
    df_rows["Clicks"]      = _distribute_clicks(scaled_imps)
    
    final_df = df_rows.drop(columns=["_weight"])
    total_clk_sum = sum(df_rows["Clicks"])
    total_row = {
        "City": "Grand Total", "Creative": "", "Impressions": total_imp,
        "Clicks": total_clk_sum, "Click Rate (CTR)": pct(total_clk_sum, total_imp)
    }
    
    info["cities_found"] = len(df_rows["City"].unique())
    return final_df.to_dict(orient="records"), total_row, info


# -- Sheet 9 - Creative --------------------------------------------------------

def build_sheet9_creative(df1: pd.DataFrame, df2: pd.DataFrame,
                          total_imp: int, total_clk: int):
    info = {"warnings": []}
    li_col1 = detect_line_item_col(df1)

    creative_col = None
    for cand in ["creative", "ad name", "ad content", "ad"]:
        for col in df2.columns:
            if cand == str(col).strip().lower():
                creative_col = col; break
        if creative_col: break

    if not creative_col:
        for col in df2.columns:
            cl = str(col).strip().lower()
            if "creative" in cl and "id" not in cl:
                creative_col = col; break

    if not creative_col:
        for cand in ["creative", "ad"]:
            for col in df2.columns:
                if cand in str(col).strip().lower():
                    creative_col = col; break
            if creative_col: break

    if not creative_col:
        creative_col = detect_line_item_col(df2)

    if not creative_col:
        return [], {"Creative": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": pct(total_clk, total_imp)}, info

    li_col2 = detect_line_item_col(df2)

    matched_df2 = pd.DataFrame()
    if li_col1 and li_col2:
        items1 = [str(v).strip() for v in df1[li_col1].dropna().unique() if str(v).strip()]

        def get_tokens(s):
            s = str(s).lower()
            # If format is "ID|Name", prioritize tokens from the Name part
            if "|" in s:
                s = s.split("|")[-1]
            return set(re.findall(r'[a-z0-9]+', s))

        def matches_any_line_item(v2_raw):
            v2_tokens = get_tokens(str(v2_raw))
            if not v2_tokens: return False
            for i1 in items1:
                i1_tokens = get_tokens(i1)
                if not i1_tokens: continue
                # Match if tokens overlap significantly or one is a subset
                if i1_tokens.issubset(v2_tokens) or v2_tokens.issubset(i1_tokens):
                    return True
            return False

        mask = df2[li_col2].apply(matches_any_line_item)
        matched_df2 = df2[mask].copy()

    if matched_df2.empty:
        info["warnings"].append("No creative-based matches found for the provided Line Items.")
        return [], {"Creative": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": pct(total_clk, total_imp)}, info

    imp_col2 = clk_col2 = None
    for col in matched_df2.columns:
        cl = col.strip().lower()
        if "impression" in cl: imp_col2 = col
        if "click" in cl: clk_col2 = col

    agg_df = (
        matched_df2
        .assign(
            _creative=matched_df2[creative_col].astype(str).str.strip(),
            _imp=matched_df2[imp_col2].apply(safe_float) if imp_col2 else 1.0,
            _clk=matched_df2[clk_col2].apply(safe_float) if clk_col2 else 1.0
        )
        .query("_creative.str.lower() not in ('nan', 'none', '', 'unknown')")
        .groupby("_creative", sort=False)[["_imp", "_clk"]]
        .sum()
        .reset_index()
    )

    unique_creatives = list(agg_df["_creative"])
    if not unique_creatives:
        return [], {"Creative": "Grand Total", "Impressions": total_imp, "Clicks": total_clk, "Click Rate (CTR)": pct(total_clk, total_imp)}, info

    imp_weights = [max(float(w), 1.0) for w in agg_df["_imp"]]
    clk_weights = [max(float(w), 0.0) for w in agg_df["_clk"]]
    w_imp_total = sum(imp_weights)
    if sum(clk_weights) == 0:
        clk_weights = imp_weights
    w_clk_total = sum(clk_weights)

    def largest_remainder(weights, w_total, total):
        if w_total == 0: w_total = 1.0
        exact = [(w / w_total) * total for w in weights]
        floored = [int(e) for e in exact]
        fracs = [e - f for e, f in zip(exact, floored)]
        remainder = int(total - sum(floored))
        order = sorted(range(len(weights)), key=lambda i: fracs[i], reverse=True)
        for i in range(remainder):
            floored[order[i]] += 1
        return floored

    crea_imps = largest_remainder(imp_weights, w_imp_total, total_imp)

    overall_ctr = total_clk / total_imp if total_imp > 0 else 0
    min_ctr = max(0, overall_ctr - 0.0014)
    max_ctr = overall_ctr + 0.0014

    raw_clks = [max(0.01, imp * rand_float(min_ctr, max_ctr)) if imp >= 180 else 0.0 for imp in crea_imps]
    w_clk_total = sum(raw_clks)
    if w_clk_total == 0:
        raw_clks = [1.0 if imp >= 180 else 0.0 for imp in crea_imps]
        w_clk_total = sum(raw_clks)

    crea_clks = largest_remainder(raw_clks, w_clk_total, total_clk)
    rows = [
        {"Creative": c, "Impressions": imp, "Clicks": clk, "Click Rate (CTR)": pct(clk, imp)}
        for c, imp, clk in zip(unique_creatives, crea_imps, crea_clks)
    ]
    rows.sort(key=lambda x: str(x["Creative"]).strip().lower())

    total_row = {
        "Creative": "Grand Total",
        "Impressions": total_imp,
        "Clicks": total_clk,
        "Click Rate (CTR)": pct(total_clk, total_imp)
    }
    return rows, total_row, info


# -- Sheet 10 - Apps / URLs ----------------------------------------------------

def build_sheet10_apps(
    app_urls_str: str,
    total_imp: int,
    total_clk: int,
    selected_sheet: str,
) -> tuple[list[dict], dict, dict]:

    info: dict = {
        "warnings": [],
        "selected_sheet": selected_sheet,
        "validated_urls": [],
        "newly_appended_urls": [],
    }

    _JUNK = {"nan", "none", "", "app/url", "app", "url", "site", "domain"}

    # Parse user-supplied URLs
    user_apps: list[str] = [
        _clean_url(u) for u in app_urls_str.split("\n")
        if _clean_url(u) not in _JUNK
    ]

    # CHANGED: validate against selected sheet only; append if not found
    all_user_urls, newly_appended = validate_and_prepare_urls(user_apps, selected_sheet)
    info["validated_urls"]      = all_user_urls
    info["newly_appended_urls"] = newly_appended

    if newly_appended:
        info["warnings"].append(
            f"{len(newly_appended)} URL(s) not found in '{selected_sheet}' — "
            "appended to that sheet: " + ", ".join(newly_appended[:5])
            + ("…" if len(newly_appended) > 5 else "")
        )

    # Remaining DB URLs from the selected sheet (excluding user-supplied ones)
    all_user_set  = set(all_user_urls)
    db_sheet_urls = [u for u in get_urls_from_multiple_sheets(selected_sheet)
                     if u not in all_user_set and u not in _JUNK]
    random.shuffle(db_sheet_urls)

    # Build filler pool: DB sheet URLs only (no pivot)
    filler_pool: list[str] = []
    seen_filler: set[str]  = set()
    for u in db_sheet_urls:
        if u not in seen_filler:
            filler_pool.append(u)
            seen_filler.add(u)

    # CHANGED: Target row count — fully random within impression-based ranges
    if total_imp < 50_000:
        target_count = random.randint(30, 60)
    elif total_imp < 100_000:
        target_count = random.randint(50, 90)
    elif total_imp < 200_000:
        target_count = random.randint(80, 120)
    else:
        target_count = random.randint(110, 150)

    real_total = len(all_user_urls) + len(filler_pool)
    if real_total > 0:
        target_count = min(target_count, real_total)

    # Place up to 15 validated user URLs spread across top 15 positions
    n_val    = len(all_user_urls)
    num_top  = min(n_val, 15)
    top_val_urls  = all_user_urls[:num_top]
    remaining_val = all_user_urls[num_top:]

    if num_top > 1:
        zone_size = 15 / num_top
        spread_indices: list[int] = []
        for i in range(num_top):
            z_start = int(i * zone_size)
            z_end   = int((i + 1) * zone_size) - 1
            prev    = spread_indices[-1] if spread_indices else -1
            spread_indices.append(random.randint(max(z_start, prev + 1), z_end))
    elif num_top == 1:
        spread_indices = [random.randint(0, 5)]
    else:
        spread_indices = []

    top_block: list[str | None] = [None] * 15
    for idx, url in zip(spread_indices, top_val_urls):
        top_block[idx] = url

    fil_idx = 0
    for i in range(15):
        if top_block[i] is None:
            if fil_idx < len(filler_pool):
                top_block[i] = filler_pool[fil_idx]
                fil_idx += 1

    all_apps: list[str] = [u for u in top_block if u is not None]

    tail_pool = remaining_val + filler_pool[fil_idx:]
    needed    = max(0, target_count - len(all_apps))
    all_apps += tail_pool[:needed]

    if not all_apps:
        all_apps = user_apps[:]
        info["warnings"].append(
            "No URLs found in DB or pivot — falling back to raw user-supplied URLs."
        )

    n = len(all_apps)

    # -- Impression Distribution: top-site scaled by total_imp --
    # Target top-site impression range based on total impressions
    if total_imp < 50_000:
        top_lo = max(1_000, int(total_imp * 0.20))
        top_hi = min(14_000, int(total_imp * 0.26))
    elif total_imp < 100_000:
        top_lo = 12_000
        top_hi = 20_000
    elif total_imp < 200_000:
        top_lo = 18_000
        top_hi = 30_000
    elif total_imp < 500_000:
        top_lo = 35_000
        top_hi = 50_000
    elif total_imp < 1_000_000:
        top_lo = 50_000
        top_hi = 65_000
    else:
        top_lo = 70_000
        top_hi = 85_000

    # Safety: never assign more than 45% of total to top site
    top_lo = min(top_lo, int(total_imp * 0.38))
    top_hi = min(top_hi, int(total_imp * 0.45))
    if top_lo >= top_hi:
        top_lo = max(1, top_hi - 500)
    top_lo = max(1, top_lo)
    target_top = random.randint(top_lo, top_hi)

    # Generate decreasing weights for positions 1..n-1
    other_weights: list[float] = []
    for i in range(1, n):
        if i < 8:
            w = random.uniform(0.55, 0.90)
        elif i < 20:
            w = random.uniform(0.08, 0.25)
        elif i < 85:
            w = random.uniform(0.003, 0.04)
        else:
            w = random.uniform(0.0005, 0.006)
        other_weights.append(w)

    # Scale other_weights so they fill (total_imp - target_top) exactly
    remainder_imp = max(0, total_imp - target_top)
    sum_oth = sum(other_weights) if other_weights else 1.0
    if sum_oth > 0 and remainder_imp > 0:
        scale = remainder_imp / sum_oth
        other_weights = [w * scale for w in other_weights]

    imp_weights: list[float] = [float(target_top)] + other_weights
    w_total = sum(imp_weights)

    def largest_remainder(weights: list[float], w_tot: float,
                          total: int, min_val: int = 1) -> list[int]:
        if not weights: return []
        if w_tot == 0: w_tot = 1.0
        exact   = [(w / w_tot) * total for w in weights]
        floored = (
            [int(round(e)) for e in exact]
            if total < len(weights) * min_val
            else [max(min_val, int(round(e))) if w > 0 else 0
                  for w, e in zip(weights, exact)]
        )
        diff = total - sum(floored)
        if diff:
            fracs = [e - int(e) for e in exact]
            order = sorted(range(len(weights)),
                           key=lambda i: fracs[i], reverse=(diff > 0))
            for i in range(abs(diff)):
                idx = order[i % len(order)]
                if diff > 0:
                    floored[idx] += 1
                elif floored[idx] > min_val:
                    floored[idx] -= 1
        return floored

    app_imps = largest_remainder(imp_weights, w_total, total_imp, min_val=1)
    app_imps = deduplicate_preserving_sum(app_imps, gap=7)

    ranked_indices = sorted(range(n), key=lambda idx: app_imps[idx], reverse=True)

    # CHANGED: CTR assignment with top 6–8 band (random each run) at 0.45–0.52%
    # All others at 0.35–0.56%, hard cap 0.56% enforced
    CTR_MIN   = 0.0035   # 0.35% floor
    CTR_MAX   = 0.0056   # 0.56% hard cap
    CTR_TOP_LO = 0.0045  # 0.45% top-band floor
    CTR_TOP_HI = 0.0052  # 0.52% top-band ceiling

    top_n = random.randint(6, 8)  # random cutoff for premium band each run

    raw_clks: list[float] = [0.0] * n
    for rank, idx in enumerate(ranked_indices, 1):
        imp = app_imps[idx]
        if imp < 180:
            continue
        if rank <= top_n:
            raw_clks[idx] = imp * random.uniform(CTR_TOP_LO, CTR_TOP_HI)
        else:
            raw_clks[idx] = imp * random.uniform(CTR_MIN, CTR_MAX)

    w_clk = sum(raw_clks) or float(n)
    app_clks = largest_remainder(raw_clks, w_clk, total_clk, min_val=0)

    for i, imp in enumerate(app_imps):
        if imp >= 180 and app_clks[i] < 1 and sum(app_clks) < total_clk:
            app_clks[i] = 1

    # Drift correction — respect CTR_MAX hard cap
    drift = total_clk - sum(app_clks)
    if drift:
        eligible = list(range(n)); random.shuffle(eligible)
        if drift > 0:
            for _ in range(drift):
                tgt = next(
                    (i for i in eligible
                     if app_imps[i] > 0 and (app_clks[i] + 1) / app_imps[i] <= CTR_MAX),
                    max(eligible, key=lambda i: app_imps[i])
                )
                app_clks[tgt] += 1
        else:
            to_remove = abs(drift)
            for idx in eligible:
                if not to_remove: break
                while app_clks[idx] > 1 and to_remove:
                    if (app_clks[idx] - 1) / app_imps[idx] >= CTR_MIN:
                        app_clks[idx] -= 1; to_remove -= 1
                    else:
                        break
            if to_remove:
                for idx in eligible:
                    if not to_remove: break
                    while app_clks[idx] > 1 and to_remove:
                        app_clks[idx] -= 1; to_remove -= 1

    rows: list[dict] = [
        {"App/URL": url, "Impressions": imp, "Clicks": clk, "Click Rate (CTR)": pct(clk, imp)}
        for url, imp, clk in zip(all_apps, app_imps, app_clks)
    ]
    rows.sort(key=lambda x: str(x["App/URL"]).strip().lower())
    rows.sort(key=lambda x: x["Impressions"], reverse=True)

    total_row: dict = {
        "App/URL": "Grand Total",
        "Impressions": total_imp,
        "Clicks": total_clk,
        "Click Rate (CTR)": pct(total_clk, total_imp),
    }
    return rows, total_row, info


def _extract_pivot_urls_internal(df1: pd.DataFrame, df2: pd.DataFrame) -> list[str]:
    """Internal helper — same logic as original extract_pivot_urls."""
    li_col1 = detect_line_item_col(df1)
    li_col2 = detect_line_item_col(df2)
    if not li_col1 or not li_col2:
        return []
    items1_set = {str(v).strip().lower() for v in df1[li_col1].dropna().unique() if str(v).strip()}
    if not items1_set:
        return []

    def is_match(v):
        v_low = str(v).lower()
        return any(i in v_low for i in items1_set)

    mask = df2[li_col2].apply(is_match)
    matched_df2 = df2[mask].copy()
    if matched_df2.empty:
        return []

    url_col = None
    for col in matched_df2.columns:
        cl = str(col).strip().lower()
        if cl in ["app/url", "url", "site", "domain", "app", "app url", "bundle id", "website", "apps"]:
            url_col = col; break
    if not url_col:
        for col in matched_df2.columns:
            cl = str(col).strip().lower()
            if any(k in cl for k in ["url", "site", "domain", "app", "bundle", "website"]):
                url_col = col; break
    if not url_col:
        for col in matched_df2.columns:
            if col != li_col2:
                url_col = col; break
    if not url_col:
        return []

    pivot_urls = (
        matched_df2[url_col].dropna().astype(str).str.strip().unique().tolist()
    )
    return [u for u in pivot_urls if u.lower() not in ["nan", "none", "", "app", "url", "site", "domain"]]


def parse_banner_file(filepath: Path) -> tuple[pd.DataFrame, int, int, str]:
    """
    Parse the banner-format Excel file.
    Structure:
      Row 1: [... , start_date, line_item_name, ..., 'Daily']
      Row 2: [... , end_date,   'Target impressions', ..., value]
      ...
      Row N: [... , 'Daily Report', total_imp, total_clk, ctr]
      Row N+1: [... , 'Date', 'Impressions', 'Clicks', 'CTR']
      Row N+2+: date data rows

    Returns: (date_df, total_imp, total_clk, line_item)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    # Line item: first row, second non-None value
    line_item = ""
    for row in all_rows:
        non_none = [v for v in row if v is not None]
        if len(non_none) >= 2:
            line_item = str(non_none[1]).strip()
            break

    # Find totals from "Daily Report" row and header from "Date" row
    total_imp = total_clk = 0
    header_row_idx = None
    for i, row in enumerate(all_rows):
        non_none = [v for v in row if v is not None]
        if not non_none:
            continue
        first = str(non_none[0]).strip()
        if first == "Daily Report" and len(non_none) >= 3:
            total_imp = safe_int(non_none[1])
            total_clk = safe_int(non_none[2])
        if first == "Date":
            header_row_idx = i
            break

    # Parse date rows
    date_rows = []
    if header_row_idx is not None:
        for row in all_rows[header_row_idx + 1:]:
            non_none = [v for v in row if v is not None]
            if len(non_none) >= 3:
                date_rows.append({
                    "Date":        str(non_none[0]).strip(),
                    "Impressions": safe_int(non_none[1]),
                    "Clicks":      safe_int(non_none[2]),
                })

    date_df = pd.DataFrame(date_rows)
    return date_df, total_imp, total_clk, line_item


def process_file(job_id: str, filepath1: Path, filepath2: Path | None,
                 app_urls: str = "", selected_sheet: str = "",
                 is_banner: bool = False, city_sheet: str = "Master Database"):
    print(f"[DEBUG] process_file thread started for Job: {job_id}")
    try:
        jobs[job_id]["status"] = "processing"

        if is_banner:
            # --- Banner format: parse the special daily-report xlsx ---
            date_df, total_imp, total_clk, line_item = parse_banner_file(filepath1)
            if total_imp == 0 and not date_df.empty:
                total_imp = int(date_df["Impressions"].apply(safe_float).sum())
                total_clk = int(date_df["Clicks"].apply(safe_float).sum())
            if date_df.empty:
                raise ValueError("Banner file contains no date data.")
            ctr_reach = pct(total_clk, total_imp)

            # Synthetic df1 with the line item — used by City & Creative builders
            df1 = pd.DataFrame({
                "Line Item Name": [line_item],
                "Impressions":    [total_imp],
                "Clicks":         [total_clk],
            })

            s2, s2t = build_sheet2_date(date_df, total_imp, total_clk, ctr_reach, is_banner=True)
        else:
            # --- Standard (video) format ---
            ext1 = filepath1.suffix.lower()
            df1  = pd.read_csv(filepath1) if ext1 == ".csv" else pd.read_excel(filepath1)
            if df1.empty:
                raise ValueError("File 1 contains no data.")
            total_imp = int(df1.get("Impressions", pd.Series([0])).apply(safe_float).sum())
            total_clk = int(df1.get("Clicks",      pd.Series([0])).apply(safe_float).sum())
            ctr_reach = pct(total_clk, total_imp)
            s2, s2t   = build_sheet2_date(df1, total_imp, total_clk, ctr_reach, is_banner=False)

        df2 = None
        if filepath2 is not None:
            ext2 = filepath2.suffix.lower()
            df2  = pd.read_csv(filepath2) if ext2 == ".csv" else pd.read_excel(filepath2)
            if df2.empty:
                df2 = None
                jobs[job_id]["warnings"] = jobs[job_id].get("warnings", []) + ["File 2 is empty - Sheet 8 skipped."]

        s1        = build_sheet1_reach(total_imp, total_clk)
        s3, s3t   = build_sheet3_timeofday(total_imp, total_clk, ctr_reach)
        s4, s4t   = build_sheet4_age(total_imp, total_clk, ctr_reach)
        s5, s5t   = build_sheet5_gender(total_imp, total_clk, ctr_reach)
        s6, s6t   = build_sheet6_device(total_imp, total_clk, ctr_reach)
        s7, s7t   = build_sheet7_exchange(total_imp, total_clk, ctr_reach)

        city_info     = {}
        creative_info = {}
        s8 = s8t = s9 = s9t = None
        if df2 is not None:
            s8, s8t, city_info     = build_sheet8_city(df1, df2, total_imp, city_sheet, is_banner=is_banner)
            s9, s9t, creative_info = build_sheet9_creative(df1, df2, total_imp, total_clk)

        output_path = OUTPUT_DIR / f"report_{job_id}.xlsx"
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "REACH"
        h1 = ["Impressions","Clicks","Click Rate (CTR)","Reach","Frequency"]
        f1 = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
        write_sheet(ws1, h1, s1, formulas=f1, alignments="center")

        ws2 = wb.create_sheet("DATE")
        if is_banner:
            h2 = ["Date","Impressions","Clicks","Click Rate (CTR)",
                  "Viewable Impressions","Measurable Impressions","Viewability"]
            f2 = {
                "Click Rate (CTR)": "{Clicks}/{Impressions}",
                "Viewability": "{Viewable Impressions}/{Measurable Impressions}",
            }
        else:
            h2 = ["Date","Impressions","Clicks","Click Rate (CTR)","Viewable Impressions",
                  "Measurable Impressions","Viewability","Sum of Starts (Video)",
                  "Sum of Complete Views (Video)","VCR (Completion Rate)"]
            f2 = {
                "Click Rate (CTR)": "{Clicks}/{Impressions}",
                "Viewability": "{Viewable Impressions}/{Measurable Impressions}",
                "VCR (Completion Rate)": "{Sum of Complete Views (Video)}/{Sum of Starts (Video)}"
            }
        t2_align = {h: "right" for h in h2}
        t2_align[h2[0]] = "left"
        write_sheet(ws2, h2, s2, s2t, formulas=f2, alignments="center", total_alignments=t2_align)

        # CHANGED: pass selected_sheet to build_sheet10_apps
        ws10 = wb.create_sheet("APP URL")
        s10, s10t, app_info = build_sheet10_apps(
            app_urls, total_imp, total_clk, selected_sheet
        )
        h10 = ["App/URL", "Impressions", "Clicks", "Click Rate (CTR)"]
        f10 = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
        align10 = {"App/URL": "left", "Impressions": "center", "Clicks": "center", "Click Rate (CTR)": "center"}
        write_sheet(ws10, h10, s10, s10t, formulas=f10, alignments=align10, total_alignments={h: "right" if i>0 else "left" for i, h in enumerate(h10)})
        jobs[job_id]["app_info"] = app_info

        ws3 = wb.create_sheet("TIME OF DAY")
        h3 = ["Time of Day","Impressions","Clicks","Click Rate (CTR)"]
        f_ctr = {"Click Rate (CTR)": "{Clicks}/{Impressions}"}
        align3 = {"Time of Day": "left", "Impressions": "center", "Clicks": "center", "Click Rate (CTR)": "center"}
        write_sheet(ws3, h3, s3, s3t, formulas=f_ctr, alignments=align3, total_alignments={h: "right" if i>0 else "left" for i, h in enumerate(h3)})

        ws7 = wb.create_sheet("EXCHANGE")
        h7 = ["Exchange","Impressions","Clicks","Click Rate (CTR)"]
        align7 = {"Exchange": "left", "Impressions": "center", "Clicks": "center", "Click Rate (CTR)": "center"}
        write_sheet(ws7, h7, s7, s7t, formulas=f_ctr, alignments=align7, total_alignments={h: "right" if i>0 else "left" for i, h in enumerate(h7)})

        ws6 = wb.create_sheet("DEVICE")
        h6 = ["Device Type","Impressions","Clicks","Click Rate (CTR)"]
        align6 = {"Device Type": "left", "Impressions": "center", "Clicks": "center", "Click Rate (CTR)": "center"}
        write_sheet(ws6, h6, s6, s6t, formulas=f_ctr, alignments=align6, total_alignments={h: "right" if i>0 else "left" for i, h in enumerate(h6)})

        if s9 is not None:
            ws9 = wb.create_sheet("CREATIVE")
            h9 = ["Creative","Impressions","Clicks","Click Rate (CTR)"]
            align9 = {"Creative": "left", "Impressions": "center", "Clicks": "center", "Click Rate (CTR)": "center"}
            write_sheet(ws9, h9, s9, s9t, formulas=f_ctr, alignments=align9, total_alignments={h: "right" if i>0 else "left" for i, h in enumerate(h9)})

        if s8 is not None:
            ws8 = wb.create_sheet("CITY")
            h8 = ["City","Impressions","Clicks","Click Rate (CTR)"]
            align8 = {"City": "left", "Impressions": "center", "Clicks": "center", "Click Rate (CTR)": "center"}
            write_sheet(ws8, h8, s8, s8t, formulas=f_ctr, alignments=align8, total_alignments={h: "right" if i>0 else "left" for i, h in enumerate(h8)})
            jobs[job_id]["city_info"] = city_info

        ws4 = wb.create_sheet("AGE")
        h4 = ["Age","Impressions","Clicks","Click Rate (CTR)"]
        align4 = {"Age": "left", "Impressions": "center", "Clicks": "center", "Click Rate (CTR)": "center"}
        write_sheet(ws4, h4, s4, s4t, formulas=f_ctr, alignments=align4, total_alignments={h: "right" if i>0 else "left" for i, h in enumerate(h4)})

        ws5 = wb.create_sheet("GENDER")
        h5 = ["Gender","Impressions","Clicks","Click Rate (CTR)"]
        align5 = {"Gender": "left", "Impressions": "center", "Clicks": "center", "Click Rate (CTR)": "center"}
        write_sheet(ws5, h5, s5, s5t, formulas=f_ctr, alignments=align5, total_alignments={h: "right" if i>0 else "left" for i, h in enumerate(h5)})

        wb.save(output_path)
        
        # Generate HTML report with the same data
        html_sheets_data = {
            'reach': {
                'title': 'Reach Overview',
                'headers': h1,
                'rows': s1,
                'total_row': None
            },
            'date': {
                'title': 'Date Breakdown',
                'headers': h2,
                'rows': s2,
                'total_row': s2t
            },
            'appurl': {
                'title': 'App & URL Performance',
                'headers': h10,
                'rows': s10,
                'total_row': s10t
            },
            'timeofday': {
                'title': 'Time of Day',
                'headers': h3,
                'rows': s3,
                'total_row': s3t
            },
            'exchange': {
                'title': 'Exchange Performance',
                'headers': h7,
                'rows': s7,
                'total_row': s7t
            },
            'device': {
                'title': 'Device Breakdown',
                'headers': h6,
                'rows': s6,
                'total_row': s6t
            }
        }
        
        if s9 is not None:
            html_sheets_data['creative'] = {
                'title': 'Creative Assets',
                'headers': h9,
                'rows': s9,
                'total_row': s9t
            }
        
        if s8 is not None:
            html_sheets_data['city'] = {
                'title': 'City Breakdown',
                'headers': h8,
                'rows': s8,
                'total_row': s8t
            }
        
        html_sheets_data.update({
            'age': {
                'title': 'Age Demographics',
                'headers': h4,
                'rows': s4,
                'total_row': s4t
            },
            'gender': {
                'title': 'Gender Demographics',
                'headers': h5,
                'rows': s5,
                'total_row': s5t
            }
        })
        
        html_output_path = generate_html_report(job_id, html_sheets_data)
        jobs[job_id]["status"] = "done"
        jobs[job_id]["output"] = str(output_path)
        jobs[job_id]["html_output"] = html_output_path

    except Exception as exc:
        jobs[job_id]["status"] = "error"
        jobs[job_id]["error"]  = str(exc)
    finally:
        for fp in [filepath1, filepath2]:
            if fp:
                try:
                    fp.unlink(missing_ok=True)
                except Exception:
                    pass


# -- Language management -------------------------------------------------------
def _norm_lang(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _load_db_master() -> pd.DataFrame:
    try:
        df = pd.read_excel(APP_DB_FILE, sheet_name=_DB_MASTER_SHEET,
                           header=_DB_HEADER_ROW - 1)
    except Exception:
        return pd.DataFrame()

    if _DB_LANG_COL not in df.columns or _DB_URL_COL not in df.columns:
        return pd.DataFrame()

    df = df[df[_DB_LANG_COL].astype(str).str.strip().str.lower()
            != _DB_LANG_COL.lower()]

    df[_DB_LANG_COL] = df[_DB_LANG_COL].ffill()

    df = df[df[_DB_URL_COL].notna()]
    df = df[~df[_DB_URL_COL].astype(str).str.strip().str.lower().isin(_DB_JUNK)]
    df = df.reset_index(drop=True)
    return df


def _load_db_records() -> list[dict]:
    df = _load_db_master()
    out: list[dict] = []
    if df.empty:
        return out
    for _, row in df.iterrows():
        lang = str(row[_DB_LANG_COL]).strip()
        url  = str(row[_DB_URL_COL]).strip()
        if not lang or not url or _clean_url(url) in _DB_JUNK:
            continue
        out.append({"language": lang, "url": url})
    return out


def list_db_languages() -> list[str]:
    seen: set[str] = set()
    order: list[str] = []
    try:
        wb = openpyxl.load_workbook(APP_DB_FILE, read_only=True, data_only=True)
        ws = wb[_DB_MASTER_SHEET] if _DB_MASTER_SHEET in wb.sheetnames else wb.active
        header_row = next(ws.iter_rows(min_row=_DB_HEADER_ROW,
                                        max_row=_DB_HEADER_ROW, values_only=True))
        try:
            lang_idx = header_row.index(_DB_LANG_COL)
        except ValueError:
            lang_idx = 2
        for row in ws.iter_rows(min_row=_DB_HEADER_ROW + 1, values_only=True):
            val = row[lang_idx] if lang_idx < len(row) else None
            if val is None:
                continue
            s = str(val).strip()
            if not s or s.lower() == _DB_LANG_COL.lower():
                continue
            key = _norm_lang(s)
            if key in seen:
                continue
            seen.add(key)
            order.append(s)
        wb.close()
    except Exception:
        pass
    return order


def list_languages_for_ui() -> list[str]:
    db_langs = list_db_languages()
    seen_norm: dict[str, str] = {}
    ordered: list[str] = []
    for name in CURATED_LANGUAGES + db_langs:
        key = _norm_lang(name)
        if key and key not in seen_norm:
            seen_norm[key] = name
            ordered.append(name)
    return ordered


def _find_language_row_range(ws, language: str) -> tuple[int, int] | None:
    target = _norm_lang(language)
    lang_col_idx = 3
    first = None
    last  = None
    current_lang = None
    for r in range(_DB_HEADER_ROW + 1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=lang_col_idx).value
        if cell_val is not None and str(cell_val).strip().lower() != _DB_LANG_COL.lower():
            current_lang = _norm_lang(cell_val)
        if current_lang == target:
            if first is None:
                first = r
            last = r
    if first is None:
        return None
    return first, last


def add_language_to_db(language: str) -> bool:
    language = str(language or "").strip()
    if not language:
        return False

    with _DB_LOCK:
        try:
            wb = openpyxl.load_workbook(APP_DB_FILE)
            ws = wb[_DB_MASTER_SHEET] if _DB_MASTER_SHEET in wb.sheetnames else wb.active
            if _find_language_row_range(ws, language) is not None:
                wb.close()
                return False
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=1, value=1)
            ws.cell(row=last_row, column=2, value="")
            ws.cell(row=last_row, column=3, value=language)
            wb.save(APP_DB_FILE)
            wb.close()
            return True
        except Exception as e:
            print(f"[DB] add_language_to_db failed for {language}: {e}")
            return False


def remove_language_from_db(language: str) -> bool:
    language = str(language or "").strip()
    if not language:
        return False

    with _DB_LOCK:
        try:
            wb = openpyxl.load_workbook(APP_DB_FILE)
            ws = wb[_DB_MASTER_SHEET] if _DB_MASTER_SHEET in wb.sheetnames else wb.active
            rng = _find_language_row_range(ws, language)
            if rng is None:
                wb.close()
                return False
            start_r, end_r = rng
            count = end_r - start_r + 1
            ws.delete_rows(start_r, count)
            wb.save(APP_DB_FILE)
            wb.close()
            return True
        except Exception as e:
            print(f"[DB] remove_language_from_db failed for {language}: {e}")
            return False


# -- HTML Report Generator ---------------------------------------------------
def generate_html_report(job_id: str, sheets_data: dict) -> str:
    """
    Generate HTML report from sheets_data.
    sheets_data = {
        'reach': {'title': str, 'rows': [row_dicts], 'headers': [str]},
        'date': {...},
        ...
    }
    Returns path to generated HTML file.
    """
    output_path = OUTPUT_DIR / f"report_{job_id}.html"
    
    # Build tabs JavaScript object
    tabs_js = {}
    for sheet_name, sheet_info in sheets_data.items():
        if not sheet_info.get('rows'):
            continue
        
        headers = sheet_info.get('headers', [])
        rows = sheet_info.get('rows', [])
        total_row = sheet_info.get('total_row', None)
        title = sheet_info.get('title', sheet_name.title())
        
        # Convert row dicts to arrays
        row_arrays = []
        for row in rows:
            row_array = [str(row.get(h, "")) for h in headers]
            row_arrays.append(row_array)
        
        # Append total row if present
        if total_row:
            total_array = [str(total_row.get(h, "")) for h in headers]
            # Ensure the first cell says "Grand Total" if it's empty
            if total_array and (not total_array[0] or total_array[0].strip() == ""):
                total_array[0] = "Grand Total"
            row_arrays.append(total_array)
        
        # Determine column alignment
        cols = []
        for h in headers:
            is_num = any(x in h.lower() for x in ['impression', 'click', 'rate', 'reach', 'frequency', 'cpm', 'cost', 'cpc', 'roas', 'revenue', 'hour', 'age'])
            cols.append({
                'label': h,
                'align': 'num' if is_num else 'left'
            })
        
        tabs_js[sheet_name] = {
            'title': title,
            'cols': cols,
            'rows': row_arrays
        }
    
    # Convert to JSON for embedding
    import json
    tabs_json = json.dumps(tabs_js)
    
    # Encode logo image so the report is self-contained
    logo_path = Path(__file__).resolve().parent / "assets" / "BILLION TAGS PNG white.png"
    logo_data = ""
    if logo_path.exists():
        try:
            logo_data = base64.b64encode(logo_path.read_bytes()).decode("ascii")
        except Exception:
            logo_data = ""
    else:
        print(f"[report] logo not found at {logo_path}")
    
    # Determine initial tab
    initial_tab = 'reach' if 'reach' in tabs_js else list(tabs_js.keys())[0] if tabs_js else 'reach'
    
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Billion Tags Report</title>
  <link rel="icon" href="https://billiontags-media.s3.ap-south-1.amazonaws.com/img/Logo/infinity.png" sizes="any">
  <link rel="apple-touch-icon" href="https://billiontags-media.s3.ap-south-1.amazonaws.com/img/Logo/infinity.png">
  <style>
*{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%;font-family:'Google Sans','Roboto',sans-serif;background:#6c27b8;color:#1d1b2c;-webkit-font-smoothing:antialiased}}
.dash{{display:flex;height:100vh;width:100%;overflow:hidden}}
.side{{width:258px;flex-shrink:0;background:#6c27b8;display:flex;flex-direction:column;height:100vh;overflow:hidden;position:relative;z-index:2}}
.brand{{padding:16px 20px 14px;display:flex;align-items:center;flex-shrink:0}}
.brand img{{height:28px;width:auto;object-fit:contain}}
.side-divider{{height:1px;background:rgba(255,255,255,.18);margin:0 20px}}
.nav-section{{flex:1;overflow-y:auto;padding:10px 10px 10px;scrollbar-width:none}}
.nav-section::-webkit-scrollbar{{display:none}}
.nav-label{{font-size:10px;font-weight:600;letter-spacing:.1em;color:rgba(255,255,255,.5);text-transform:uppercase;padding:10px 14px 6px}}
.nav-item{{display:flex;align-items:center;padding:10px 14px;border-radius:9px;color:rgba(255,255,255,.82);font-size:13px;font-weight:400;cursor:pointer;width:100%;background:transparent;border:none;text-align:left;justify-content:space-between;line-height:1;transition:background .15s,color .15s;margin-bottom:1px;font-family:inherit;white-space:nowrap}}
.nav-item:hover{{background:rgba(255,255,255,.1);color:#fff}}
.nav-item.active{{background:#fff;color:#1d1b2c;font-weight:600}}
.nav-item .n-left{{display:flex;align-items:center;gap:10px;flex:1;min-width:0}}
.nav-item .n-icon{{font-size:17px;color:rgba(255,255,255,.85);width:20px;flex-shrink:0;display:flex;align-items:center;justify-content:center}}
.nav-item.active .n-icon{{color:#6c27b8}}
.nav-item .n-label{{flex:1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.nav-item .n-badge{{font-size:10px;background:rgba(255,255,255,.2);color:rgba(255,255,255,.9);padding:2px 8px;border-radius:10px;font-weight:500;line-height:1.5;flex-shrink:0}}
.nav-item.active .n-badge{{background:rgba(108,39,184,.15);color:#6c27b8}}
.side-foot{{padding:14px 16px;border-top:1px solid rgba(255,255,255,.15);display:flex;align-items:center;gap:10px;flex-shrink:0}}
.s-avatar{{width:36px;height:36px;border-radius:50%;background:linear-gradient(135deg,#f5a87f,#d85a30);display:flex;align-items:center;justify-content:center;color:#fff;font-size:12px;font-weight:700;flex-shrink:0}}
.s-uname{{color:#f0ecff;font-size:12.5px;font-weight:600;line-height:1.3}}
.s-urole{{color:rgba(255,255,255,.55);font-size:11px;margin-top:1px}}
.main{{flex:1;min-width:0;background:#f0eff5;border-radius:28px 0 0 28px;display:flex;flex-direction:column;height:100vh;overflow:hidden;position:relative;box-shadow:-4px 0 24px rgba(0,0,0,.18)}}
.topbar{{background:#fff;border-radius:28px 0 0 0;padding:0 28px;display:flex;align-items:center;justify-content:center;height:56px;flex-shrink:0;border-bottom:1px solid #e8e5f0;position:relative}}
.topbar-title{{font-size:16px;font-weight:600;color:#1d1b2c;letter-spacing:-.01em;white-space:nowrap}}
.hamburger{{display:none;position:absolute;left:16px;top:50%;transform:translateY(-50%);background:none;border:none;cursor:pointer;padding:6px;color:#6c27b8;font-size:22px;line-height:1}}
.mobile-nav-strip{{display:none}}
.content{{flex:1;display:flex;flex-direction:column;padding:16px 20px 0;gap:12px;overflow:hidden;min-height:0}}
.toolbar{{display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap;flex-shrink:0}}
.search-box{{display:flex;align-items:center;gap:8px;background:#fff;border:1px solid #e2e0ea;border-radius:8px;padding:8px 13px;font-size:13px;color:#6b6a7a;min-width:180px;flex:1;max-width:340px}}
.search-box i{{font-size:14px;flex-shrink:0;color:#999}}
.search-box input{{flex:1;border:none;background:transparent;outline:none;font-size:13px;color:#1d1b2c;font-family:inherit;min-width:0}}
.search-box input::placeholder{{color:#aaa}}
.tools{{display:flex;gap:7px;flex-wrap:wrap}}
.tool-btn{{display:inline-flex;align-items:center;gap:5px;font-size:12.5px;background:#fff;border:1px solid #e2e0ea;border-radius:7px;padding:8px 15px;color:#5a5868;cursor:pointer;font-family:inherit;font-weight:500;transition:all .15s;line-height:1;white-space:nowrap}}
.tool-btn:hover{{background:#f4f3f8;color:#1d1b2c;border-color:#c8c4d8}}
.tool-btn.primary{{background:#6c27b8;color:#fff;border-color:#6c27b8;box-shadow:0 1px 4px rgba(108,39,184,.3)}}
.tool-btn.primary:hover{{background:#5a1da8}}
.tool-btn i{{font-size:13px}}
.table-shell{{border:1px solid #dddae8;border-radius:12px;overflow:hidden;background:#fff;box-shadow:0 1px 4px rgba(0,0,0,.04);display:flex;flex-direction:column;flex:1;min-height:0}}
.table-scroll{{overflow:auto;flex:1;min-height:0;-webkit-overflow-scrolling:touch}}
.table-scroll::-webkit-scrollbar{{height:5px;width:5px}}
.table-scroll::-webkit-scrollbar-track{{background:#f4f3f8}}
.table-scroll::-webkit-scrollbar-thumb{{background:#c8c4d8;border-radius:4px}}
table{{width:100%;border-collapse:collapse;font-size:13px;min-width:480px}}
thead{{background:#f6f5fb;position:sticky;top:0;z-index:2}}
th{{text-align:left;padding:12px 16px;font-size:11px;font-weight:600;color:#888;white-space:nowrap;text-transform:uppercase;letter-spacing:.07em;border-bottom:1px solid #e8e5f0;background:#f6f5fb;vertical-align:middle}}
th.num{{text-align:right}}
td{{padding:13px 16px;color:#1d1b2c;border-bottom:1px solid #f0eff5;white-space:nowrap;font-weight:400;background:#fff;vertical-align:middle;line-height:1.4;font-size:13px}}
tbody tr:last-child td{{border-bottom:none}}
td.num{{text-align:right;font-weight:500;color:#3a3848}}
td.primary{{font-weight:500;color:#1d1b2c}}
tbody tr.total-row td{{background:#eef3ff;color:#1d1b2c;font-weight:700}}
tbody tr{{transition:background .1s}}
tbody tr:hover td{{background:#faf9fd}}
.tbl-foot{{padding:10px 16px;border-top:1px solid #e8e5f0;display:flex;align-items:center;justify-content:space-between;font-size:12px;color:#888;background:#faf9fd;flex-wrap:wrap;gap:8px;font-weight:500;flex-shrink:0}}
.dash-footer{{padding:10px 20px;text-align:center;font-size:11.5px;color:#aaa;flex-shrink:0}}
.dash-footer a{{color:#6c27b8;text-decoration:none}}
@media(max-width:1023px) and (min-width:768px){{
  .side{{width:62px}}
  .brand{{padding:16px 0;justify-content:center}}
  .brand img{{height:20px;max-width:46px}}
  .main{{border-radius:20px 0 0 20px}}
  .topbar{{border-radius:20px 0 0 0;padding:0 20px}}
  .content{{padding:14px 16px 0}}
}}
@media(max-width:767px){{
  .side{{display:none}}
  .dash{{flex-direction:column;height:100vh}}
  .main{{border-radius:0;box-shadow:none;flex:1}}
  .topbar{{border-radius:0;height:52px;padding:0 16px}}
  .hamburger{{display:block}}
  .content{{flex:1;padding:10px 12px 0;gap:9px}}
  .toolbar{{flex-direction:column;align-items:stretch;gap:8px}}
  table{{min-width:460px}}
}}
</style>
</head>
<body>
<div class="dash">
  <aside class="side">
    <div class="brand">
      <img src="data:image/png;base64,{logo_data}" alt="Billion Tags" style="height:28px;width:auto;object-fit:contain;filter:none;">
    </div>
    <div class="side-divider"></div>
    <div class="nav-section">
      <div class="nav-label">Analytics Dimensions</div>
"""
    
    # Add nav items dynamically
    sheet_icons = {
        'reach': 'ti-broadcast',
        'date': 'ti-calendar-event',
        'appurl': 'ti-world',
        'timeofday': 'ti-clock-hour-9',
        'exchange': 'ti-arrows-exchange-2',
        'device': 'ti-devices',
        'creative': 'ti-palette',
        'city': 'ti-map-pin',
        'age': 'ti-users',
        'gender': 'ti-gender-bigender'
    }
    
    sheet_labels = {
        'reach': 'Reach',
        'date': 'Date',
        'appurl': 'App URL',
        'timeofday': 'Time of day',
        'exchange': 'Exchange',
        'device': 'Device',
        'creative': 'Creative',
        'city': 'City',
        'age': 'Age',
        'gender': 'Gender'
    }
    
    for sheet_name in tabs_js.keys():
        icon = sheet_icons.get(sheet_name, 'ti-circle')
        label = sheet_labels.get(sheet_name, sheet_name.title())
        row_count = len(tabs_js[sheet_name]['rows'])
        active_class = 'active' if sheet_name == initial_tab else ''
        html_content += f"""      <button class="nav-item {active_class}" data-tab="{sheet_name}" onclick="setTab('{sheet_name}')">
        <span class="n-left"><i class="ti {icon} n-icon"></i><span class="n-label">{label}</span></span>
        <span class="n-badge">{row_count}</span>
      </button>\n"""
    
    html_content += f"""    </div>
    <div class="side-foot">
      <div class="s-avatar">BT</div>
      <div><div class="s-uname">Report</div><div class="s-urole">Generated</div></div>
    </div>
  </aside>

  <main class="main">
    <div class="topbar">
      <button class="hamburger" onclick="openDrawer()" aria-label="Menu"><i class="ti ti-menu-2"></i></button>
      <div class="topbar-title" id="topbar-title">Report Dashboard</div>
    </div>

    <div class="content">
      <div class="toolbar">
        <div class="search-box">
          <i class="ti ti-search"></i>
          <input type="text" placeholder="Search rows…" id="search-input" oninput="doSearch(this.value)">
        </div>
      </div>

      <div class="table-shell">
        <div class="table-scroll">
          <table>
            <thead id="table-head"></thead>
            <tbody id="table-body"></tbody>
          </table>
        </div>
        <div class="tbl-foot">
          <span id="row-info">Loading...</span>
        </div>
      </div>

      <div class="dash-footer">@2026 <a href="#">Billiontags Creations Private Limited</a> | All Rights Reserved</div>
    </div>
  </main>
</div>

<script>
const tabs = {tabs_json};

function renderTable(key){{
  const d = tabs[key];
  const head = document.getElementById('table-head');
  const body = document.getElementById('table-body');
  if (!d || !d.cols || !d.rows || d.rows.length === 0) {{
    document.getElementById('topbar-title').textContent = d ? d.title : 'Report';
    head.innerHTML = '';
    body.innerHTML = '<tr><td colspan="4" style="padding:24px;text-align:center;color:#6b6a7a;">No report data available.</td></tr>';
    document.getElementById('row-info').textContent = 'No data available';
    return;
  }}
  document.getElementById('topbar-title').textContent = d.title;
  document.getElementById('search-input').value = '';
  head.innerHTML = '<tr>' + d.cols.map(c => {{
    const cls = c.align === 'num' ? 'num' : '';
    return '<th class="' + cls + '"><span class="th-inner">' + c.label + '</span></th>';
  }}).join('') + '</tr>';
  body.innerHTML = d.rows.map(r => {{
    const isTotal = /grand total|total/i.test(String(r[0] || ''));
    const rowClass = isTotal ? ' total-row' : '';
    return '<tr class="' + rowClass.trim() + '">' + r.map((cell, i) => {{
      const c = d.cols[i];
      let cls = c.align === 'num' ? 'num' : '';
      if (i === 0) cls += ' primary';
      return '<td class="' + cls.trim() + '">' + cell + '</td>';
    }}).join('') + '</tr>';
  }}).join('');
  document.getElementById('row-info').textContent = 'Showing ' + d.rows.length + ' of ' + d.rows.length + ' rows · ' + d.cols.length + ' columns';
}}

function doSearch(q){{
  q = q.toLowerCase().trim();
  const rows = document.querySelectorAll('#table-body tr');
  let visible = 0;
  rows.forEach(tr => {{
    const match = tr.textContent.toLowerCase().includes(q);
    tr.style.display = match ? '' : 'none';
    if (match) visible++;
  }});
  const cols = document.querySelectorAll('#table-head th').length;
  document.getElementById('row-info').textContent = 'Showing ' + visible + ' of ' + rows.length + ' rows · ' + cols + ' columns';
}}

function setTab(tab){{
  document.querySelectorAll('[data-tab]').forEach(b => b.classList.toggle('active', b.dataset.tab === tab));
  renderTable(tab);
}}

document.addEventListener('DOMContentLoaded', function(){{
  setTab('{initial_tab}');
}});
</script>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    return str(output_path)


# -- Routes --------------------------------------------------------------------
def cleanup_expired_jobs():
    now = time.time()
    expired = [job_id for job_id, job in jobs.items() if now - job.get("created_at", now) > MAX_REPORT_AGE]
    for job_id in expired:
        job = jobs.pop(job_id, None)
        if not job:
            continue
        try:
            Path(job.get("output", "")).unlink(missing_ok=True)
            Path(job.get("html_output", "")).unlink(missing_ok=True)
        except Exception as e:
            print(f"[cleanup] failed to remove expired job {job_id}: {e}")


@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response


# Expose App/URL DB sheet names for language dropdown
@app.route("/db-sheets", methods=["GET"])
def db_sheets():
    names = get_db_sheet_names()
    return jsonify({"sheets": names})


# Expose City DB sheet names for city dropdown
@app.route("/city-sheets", methods=["GET"])
def city_sheets_route():
    names = get_city_db_sheet_names()
    return jsonify({"sheets": names})


def build_sheet8_city(matched_df1: pd.DataFrame, matched_df2: pd.DataFrame,
                      total_imp: int, city_db_sheet: str, is_banner: bool = False):
    """
    Deterministic city breakdown.
    Keeps all unique cities from the filtered source rows.
    City DB data is only used to enrich creatives when available.
    """
    info = {"warnings": [], "debug": {}, "city_source": ""}

    if matched_df1.empty or matched_df2.empty:
        return [], None, info

    creative_col1 = next(
        (c for c in matched_df1.columns if "creative" in c.lower() and "id" not in c.lower()),
        matched_df1.columns[0],
    )

    city_col = None
    for col in matched_df2.columns:
        if str(col).strip().lower() == "city":
            city_col = col
            break
    if city_col is None:
        for col in matched_df2.columns:
            if "city" in str(col).strip().lower():
                city_col = col
                break

    if city_col is None:
        info["warnings"].append("No 'City' column detected in File 2.")
        return [], None, info

    weight_col2 = next(
        (c for c in matched_df2.columns if "impression" in c.lower() or "weight" in c.lower()),
        None,
    )
    if weight_col2 is None:
        numeric_candidates = [
            c for c in matched_df2.columns
            if pd.api.types.is_numeric_dtype(matched_df2[c])
        ]
        weight_col2 = numeric_candidates[0] if numeric_candidates else matched_df2.columns[-1]

    li_col2 = next((c for c in matched_df2.columns if "line item" in c.lower()), matched_df2.columns[0])
    li_col1 = next((c for c in matched_df1.columns if "line item" in c.lower()), matched_df1.columns[0])

    db_cities_raw = load_city_db_sheet(city_db_sheet)
    if not db_cities_raw:
        db_cities_raw = load_city_db_sheet("Master Database")
        if db_cities_raw:
            info["warnings"].append(f"Sheet '{city_db_sheet}' not found. Using Master Database.")
    db_lookup = {c["name"].strip().lower(): c for c in db_cities_raw}

    target_li_names = {
        str(v).strip()
        for v in matched_df1[li_col1].dropna().astype(str).unique()
        if str(v).strip()
    }

    def normalize_line_item(val: object) -> str:
        val_str = str(val).strip()
        if "|" in val_str:
            val_str = val_str.split("|", 1)[1].strip()
        return val_str

    filtered_df2 = matched_df2[matched_df2[li_col2].apply(lambda v: normalize_line_item(v) in target_li_names)].copy()
    info["debug"]["target_line_items"] = sorted(target_li_names)
    info["debug"]["filtered_rows"] = len(filtered_df2)

    if filtered_df2.empty:
        info["warnings"].append("No exact Line Item matches found after normalization. Using all data as fallback.")
        filtered_df2 = matched_df2.copy()

    city_groups = filtered_df2.groupby(city_col, dropna=False).agg({
        weight_col2: "sum",
        li_col2: lambda x: sorted({
            normalize_line_item(v)
            for v in x.dropna().astype(str)
            if normalize_line_item(v)
        }),
    }).reset_index()

    input_cities_map = {}
    for _, row in city_groups.iterrows():
        name = str(row[city_col]).strip()
        if name and name.lower() not in {"nan", "none", ""}:
            input_cities_map[name.lower()] = {
                "name": name,
                "weight": safe_float(row[weight_col2]),
                "creatives": row[li_col2],
            }

    final_cities = []
    for city_key, inp in input_cities_map.items():
        db_city = db_lookup.get(city_key)
        final_cities.append({
            "name": inp["name"],
            "weight": inp["weight"],
            "creatives": db_city["creatives"] if db_city and db_city["creatives"] else inp["creatives"],
        })

    final_cities.sort(key=lambda x: x["weight"], reverse=True)

    all_f1_creatives = sorted(list(set(matched_df1[creative_col1].dropna().astype(str))))
    if not all_f1_creatives:
        all_f1_creatives = ["Default Creative"]

    rows = []
    for city_info in final_cities:
        city_name = city_info["name"]
        city_weight = max(city_info["weight"], 1.0)
        city_creatives = city_info["creatives"] or all_f1_creatives
        weight_per_cr = city_weight / len(city_creatives)
        for cr in city_creatives:
            rows.append({
                "City": city_name,
                "Creative": cr,
                "_weight": weight_per_cr,
            })

    if not rows:
        for city_name in ["Sydney", "Melbourne", "Brisbane", "Perth"]:
            rows.append({"City": city_name, "Creative": all_f1_creatives[0], "_weight": 1.0})

    df_rows = pd.DataFrame(rows)
    total_weight = df_rows["_weight"].sum()
    scaled_imps = _largest_remainder(df_rows["_weight"].tolist(), total_weight, total_imp)

    df_rows["Impressions"] = scaled_imps
    df_rows["Clicks"] = _distribute_clicks(scaled_imps)

    final_df = df_rows.drop(columns=["_weight"])
    total_clk_sum = int(df_rows["Clicks"].sum())
    total_row = {
        "City": "Grand Total",
        "Creative": "",
        "Impressions": total_imp,
        "Clicks": total_clk_sum,
        "Click Rate (CTR)": pct(total_clk_sum, total_imp),
    }

    info["cities_found"] = len(final_df["City"].drop_duplicates())
    info["city_source"] = "input"
    return final_df.to_dict(orient="records"), total_row, info


@app.route("/upload", methods=["POST", "OPTIONS"])
def upload():
    print(f"\n[DEBUG] --- New upload request received ---")
    if request.method == "OPTIONS":
        return jsonify({}), 200

    if "file" not in request.files:
        return jsonify({"error": "No file in request"}), 400

    file1 = request.files["file"]
    if not file1.filename:
        return jsonify({"error": "No file selected"}), 400

    if Path(file1.filename).suffix.lower() not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Only .csv, .xlsx, .xls files allowed"}), 400

    job_id     = uuid.uuid4().hex
    save_path1 = UPLOAD_DIR / f"{job_id}_1_{secure_filename(file1.filename)}"
    file1.save(str(save_path1))

    save_path2 = None
    if "file2" in request.files and request.files["file2"].filename:
        file2 = request.files["file2"]
        if Path(file2.filename).suffix.lower() not in ALLOWED_EXTENSIONS:
            return jsonify({"error": "File 2: only .csv, .xlsx, .xls files allowed"}), 400
        save_path2 = UPLOAD_DIR / f"{job_id}_2_{secure_filename(file2.filename)}"
        file2.save(str(save_path2))

    cleanup_expired_jobs()
    jobs[job_id] = {
        "status": "queued",
        "output": None,
        "html_output": None,
        "error": None,
        "city_info": None,
        "app_info": None,
        "created_at": time.time(),
    }
    app_urls       = request.form.get("app_urls", "")
    selected_sheet = request.form.get("selected_sheet", "")
    mode           = request.form.get("mode", "video")
    is_banner      = (mode == "banner")
    city_sheet     = request.form.get("city_sheet", "Master Database")

    threading.Thread(
        target=process_file,
        args=(job_id, save_path1, save_path2, app_urls, selected_sheet, is_banner, city_sheet),
        daemon=True
    ).start()

    return jsonify({"job_id": job_id}), 202


@app.route("/status/<job_id>")
def status(job_id):
    cleanup_expired_jobs()
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Unknown job"}), 404
    base_url = request.host_url.rstrip("/")
    return jsonify({
        "status":           job["status"],
        "error":            job.get("error"),
        "city_info":        job.get("city_info"),
        "app_info":         job.get("app_info"),
        "download_url":     f"{base_url}/download/{job_id}" if job["status"] == "done" else None,
        "download_html_url":f"{base_url}/download-html/{job_id}" if job["status"] == "done" else None,
        "view_url":         f"{base_url}/report/{job_id}" if job["status"] == "done" else None,
    })


@app.route("/debug/<job_id>")
def debug_job(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Unknown job"}), 404
    return jsonify({
        "status":      job["status"],
        "error":       job.get("error"),
        "html_output": job.get("html_output"),
        "city_info":   job.get("city_info"),
        "app_info":    job.get("app_info"),
    })


@app.route("/download/<job_id>")
def download(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Unknown job"}), 404
    if job["status"] != "done":
        return jsonify({"error": "Report not ready"}), 400

    output_path = Path(job["output"])
    if not output_path.exists():
        return jsonify({"error": "Output file missing"}), 500

    return send_file(output_path, as_attachment=True,
                     download_name=f"campaign_report_{job_id[:8]}.xlsx")


@app.route("/download-html/<job_id>")
def download_html(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Unknown job"}), 404
    if job["status"] != "done":
        return jsonify({"error": "Report not ready"}), 400

    html_path = Path(job.get("html_output", ""))
    if not html_path.exists():
        return jsonify({"error": "HTML report missing"}), 500

    return send_file(html_path, as_attachment=True,
                     download_name=f"campaign_report_{job_id[:8]}.html")


@app.route("/report/<job_id>")
def report_view(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Unknown job"}), 404
    if job["status"] != "done":
        return jsonify({"error": "Report not ready"}), 400

    html_path = Path(job.get("html_output", ""))
    if not html_path.exists():
        return jsonify({"error": "HTML report missing"}), 500

    return send_file(html_path, mimetype="text/html", as_attachment=False)


# -- Language management -------------------------------------------------------
@app.route("/languages", methods=["GET", "OPTIONS"])
def list_languages_route():
    if request.method == "OPTIONS":
        return jsonify({}), 200
    return jsonify({"languages": list_languages_for_ui()})


@app.route("/languages", methods=["POST", "OPTIONS"])
def add_language_route():
    if request.method == "OPTIONS":
        return jsonify({}), 200
    data = request.get_json(silent=True)
    if isinstance(data, dict):
        language = str(data.get("language", "")).strip()
    elif isinstance(data, str):
        language = data.strip()
    else:
        language = str(request.form.get("language", "")).strip()
    if not language:
        return jsonify({"error": "language required"}), 400
    added = add_language_to_db(language)
    return jsonify({
        "ok": True,
        "added_to_db": added,
        "languages": list_languages_for_ui(),
    })


@app.route("/languages/<path:language>", methods=["DELETE", "OPTIONS"])
def remove_language_route(language: str):
    if request.method == "OPTIONS":
        return jsonify({}), 200
    removed = remove_language_from_db(language)
    return jsonify({
        "ok": True,
        "removed_from_db": removed,
        "languages": list_languages_for_ui(),
    })


@app.route("/")
def index():
    frontend_dir = BASE_DIR.parent / "Final_Report_Frontend"
    return send_file(frontend_dir / "index.html")

# -- Entry point ---------------------------------------------------------------
if __name__ == "__main__":
    print("-  Ad Campaign Report Server running at http://localhost:5000")
    print("    Open http://localhost:5000 in your browser (do NOT open Index.html directly).")
    app.run(debug=True, port=5000)
